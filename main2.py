from fastapi.security import OAuth2PasswordBearer  
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient, AsyncIOMotorGridFSBucket
from pydantic import BaseModel
from datetime import datetime, timedelta
import calendar
from jose import JWTError, jwt
from passlib.context import CryptContext
from typing import Optional
from dotenv import load_dotenv
import os
import io
from bson import ObjectId
from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File, Form, Body, Request
from starlette.requests import Request
from fastapi.responses import JSONResponse
import json

# Load env vars
load_dotenv()

MONGO_URI = os.getenv("MONGO_URI")
MONGO_DB = os.getenv("MONGO_DB")
JWT_SECRET = os.getenv("JWT_SECRET")
JWT_ALGORITHM = os.getenv("JWT_ALGORITHM")
JWT_EXPIRE_MINUTES = 14400

pwd_context = CryptContext(schemes=["pbkdf2_sha256"], deprecated="auto")

class SafeJSONResponse(JSONResponse):
    def render(self, content) -> bytes:
        def sanitize(obj):
            if isinstance(obj, float):
                if math.isnan(obj) or math.isinf(obj):
                    return 0.0
                return obj
            if isinstance(obj, dict):
                return {k: sanitize(v) for k, v in obj.items()}
            if isinstance(obj, list):
                return [sanitize(i) for i in obj]
            return obj
        return json.dumps(sanitize(content), ensure_ascii=False).encode("utf-8")

# ---------- FastAPI app ----------
app = FastAPI(default_response_class=SafeJSONResponse)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------- Mongo Connection ----------
client = AsyncIOMotorClient(MONGO_URI)
db = client[MONGO_DB]
user_collection = db["user"]   


# ---------- GridFS Helpers ----------
async def upload_to_gridfs(file_content: bytes, filename: str) -> str:
    """Upload a file to GridFS and return the file_id as string."""
    bucket = AsyncIOMotorGridFSBucket(db)
    file_id = await bucket.upload_from_stream(filename, io.BytesIO(file_content))
    return str(file_id)


async def delete_from_gridfs(file_id: str):
    """Delete a file from GridFS by its file_id string."""
    try:
        bucket = AsyncIOMotorGridFSBucket(db)
        await bucket.delete(ObjectId(file_id))
    except Exception:
        pass


def is_gridfs_id(value: str) -> bool:
    """Check if a string is a valid GridFS ObjectId (24 hex chars)."""
    if not value:
        return False
    try:
        ObjectId(value)
        return len(value) == 24
    except Exception:
        return False


# ---------- Models ----------
class UserCreate(BaseModel):
    employee_code: str
    password: str


class UserLogin(BaseModel):
    employee_code: str
    password: str


class Token(BaseModel):
    access_token: str
    token_type: str


class TokenData(BaseModel):
    employee_code: Optional[str] = None


import math

def safe_float(val) -> float:
    try:
        if val is None or str(val).strip().lower() in ("", "nan", "inf", "-inf"):
            return 0.0
        f = float(val)
        if math.isnan(f) or math.isinf(f):
            return 0.0
        return f
    except (ValueError, TypeError):
        return 0.0


# ---------- Utility functions ----------
def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)


def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=JWT_EXPIRE_MINUTES))
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)
    return encoded_jwt


async def get_user_by_employee_code(employee_code: str):
    return await user_collection.find_one({"employee_code": employee_code})


# ---------- JWT dependency ----------
from fastapi.security import OAuth2PasswordBearer

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="api/login")


async def get_current_user(token: str = Depends(oauth2_scheme)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        employee_code: str = payload.get("sub")
        if employee_code is None:
            raise credentials_exception
        token_data = TokenData(employee_code=employee_code)
    except JWTError:
        raise credentials_exception

    user = await get_user_by_employee_code(token_data.employee_code)
    if user is None:
        raise credentials_exception
    return user


async def is_valid_employee(employee_code: str) -> bool:
    emp = await db["Employee_details"].find_one({"Employee_ID": employee_code})
    return emp is not None


# ---------- PDF Serve Endpoint ----------
@app.get("/api/ope/pdf/{file_id}")
async def serve_pdf(file_id: str, current_user=Depends(get_current_user)):
    """
    Serve a PDF stored in GridFS by its file_id.
    Requires authentication.
    """
    try:
        bucket = AsyncIOMotorGridFSBucket(db)
        stream = await bucket.open_download_stream(ObjectId(file_id))
        content = await stream.read()
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"inline; filename=ticket_{file_id}.pdf",
                "Cache-Control": "private, max-age=3600"
            }
        )
    except Exception as e:
        print(f"❌ PDF serve error for {file_id}: {str(e)}")
        raise HTTPException(status_code=404, detail="PDF not found")


# ---------- Auth endpoints ----------

@app.post("/api/register")
async def register(user: UserCreate):
    print("📌 Incoming register data:", user.employee_code)

    existing = await user_collection.find_one({"employee_code": user.employee_code})
    print("📌 Existing user:", existing)

    if existing:
        raise HTTPException(
            status_code=400,
            detail="Employee already registered"
        )

    hashed_password = get_password_hash(user.password)
    print("📌 Password hashed")

    doc = {
        "employee_code": user.employee_code,
        "password_hash": hashed_password,
        "created_at": datetime.utcnow()
    }

    result = await user_collection.insert_one(doc)
    print("📌 Insert result:", result.inserted_id)

    return {"message": "Registered successfully"}

@app.post("/api/login", response_model=Token)
async def login(user: UserLogin):
    db_user = await get_user_by_employee_code(user.employee_code)
    if not db_user:
        raise HTTPException(status_code=401, detail="Incorrect employee code or password")

    if not verify_password(user.password, db_user["password_hash"]):
        raise HTTPException(status_code=401, detail="Incorrect employee code or password")

    access_token = create_access_token(
        data={"sub": user.employee_code}
    )

    return {"access_token": access_token, "token_type": "bearer"}

     
# ---------- Protected example route ----------
@app.get("/api/me")
async def read_current_user(current_user=Depends(get_current_user)):
    return {
        "employee_code": current_user["employee_code"],
        "created_at": current_user.get("created_at"),
    }


# Employee Details Fetch from Backend
@app.get("/api/employee/{employee_code}")
async def get_employee_details(employee_code: str, current_user=Depends(get_current_user)):

    emp = await db["Employee_details"].find_one({"EmpID": employee_code})

    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    return {
        "employee_id": emp.get("EmpID"),
        "employee_name": emp.get("Emp Name"),
        "designation": emp.get("Designation Name"),
        "gender": emp.get("Gender"),
        "partner": emp.get("Partner"),
        "reporting_manager_code": emp.get("ReportingEmpCode"),
        "reporting_manager_name": emp.get("ReportingEmpName"),
        "ope_limit": emp.get("OPE Limit")
    }


# ---------- OPE Data Submission ----------
@app.post("/api/ope/submit")
async def submit_ope_entry(
    date: str = Form(...),
    client: str = Form(...),
    project_id: str = Form(...),
    project_name: str = Form(...),
    project_type: str = Form(...),
    location_from: str = Form(...),
    location_to: str = Form(...),
    travel_mode: str = Form(...),
    amount: float = Form(...),
    remarks: str = Form("NA"),
    month_range: str = Form(...),
    ticket_pdf: UploadFile = File(None),
    current_user: dict = Depends(get_current_user)
):
    try:
        emp_code = current_user.get("employee_code")
        
        print(f"\n{'='*60}")
        print(f"📝 OPE SUBMISSION REQUEST")
        print(f"Submitter: {emp_code}")
        print(f"Amount: ₹{amount}")
        print(f"Month: {month_range}")
        print(f"{'='*60}\n")
        
        # ============================================
        # ✅ STEP 1: GET EMPLOYEE DETAILS
        # ============================================
        employee = await db["Employee"].find_one({"EmployeeId": emp_code})
        
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        
        reporting_manager_code = employee.get("ReportingEmpCode")
        partner_code = employee.get("PartnerEmpCode")
        reporting_manager_name = employee.get("Reporting_Manager", "Manager")
        partner_name = employee.get("Partner", "Partner")
        ope_limit = employee.get("OPE_limit", 5000)
        employee_name = employee.get("EmployeeName")
        
        print(f"👤 Employee Details:")
        print(f"   Name: {employee_name}")
        print(f"   Employee ID: {emp_code}")
        print(f"   Reporting Manager: {reporting_manager_name} ({reporting_manager_code})")
        print(f"   Partner: {partner_name} ({partner_code})")
        print(f"   OPE Limit: ₹{ope_limit}")
        
        # ============================================
        # ✅ STEP 2: CHECK COLLECTIONS
        # ============================================
        
        is_rm_in_collection = await db["Reporting_Managers"].find_one({"EmployeeId": emp_code})
        is_partner_in_collection = await db["Partner"].find_one({"PartnerEmpCode": emp_code})
        
        print(f"\n🔍 Collection Checks:")
        print(f"   Is in Reporting_Managers: {bool(is_rm_in_collection)}")
        print(f"   Is in Partner: {bool(is_partner_in_collection)}")
        print(f"   ReportingEmpCode == PartnerEmpCode: {reporting_manager_code == partner_code}")
        
        # ============================================
        # ✅ STEP 3: HANDLE PDF - GridFS Upload
        # ============================================
        ticket_pdf_id = None
        if ticket_pdf:
            pdf_content = await ticket_pdf.read()
            pdf_filename = f"ope_{emp_code}_{datetime.utcnow().timestamp()}.pdf"
            ticket_pdf_id = await upload_to_gridfs(pdf_content, pdf_filename)
            print(f"✅ PDF uploaded to GridFS: {ticket_pdf_id}")
        
        # ============================================
        # ✅ STEP 4: CREATE OPE ENTRY
        # ============================================
        entry_data = {
            "employee_id": emp_code,
            "employee_name": employee_name,
            "date": date,
            "client": client,
            "project_id": project_id,
            "project_name": project_name,
            "project_type": project_type,
            "location_from": location_from,
            "location_to": location_to,
            "travel_mode": travel_mode,
            "amount": amount,
            "remarks": remarks,
            "month_range": month_range,
            "ticket_pdf": ticket_pdf_id,       # GridFS file_id or None
            "status": "pending",
            "submission_date": datetime.utcnow(),
            "L1_approved": {"status": False},
            "L2_approved": {"status": False}
        }
        
        result = await db["OPE_data"].insert_one(entry_data)
        entry_id = str(result.inserted_id)
        
        print(f"✅ Entry created: {entry_id}")
        
        # ============================================
        # ✅ STEP 5: ROUTING LOGIC - DECISION TREE
        # ============================================
        
        if is_rm_in_collection:
            print(f"\n{'🔥'*30}")
            print(f"🔥 REPORTING MANAGER SELF-SUBMISSION")
            print(f"{'🔥'*30}")
            print(f"   RM Code: {emp_code}")
            print(f"   RM Name: {employee_name}")
            
            total_levels = 2
            
            if reporting_manager_code == partner_code:
                print(f"\n   📌 SCENARIO 1: RM Reports to Self (Partner Role)")
                
                partner_entry = await db["Partner"].find_one({"PartnerEmpCode": partner_code})
                
                if not partner_entry:
                    raise HTTPException(
                        status_code=500,
                        detail=f"Partner {partner_code} not found in Partner collection"
                    )
                
                l1_approver_code = partner_code
                l1_approver_name = partner_entry.get("Partner_Name", partner_name)
                pending_queue_code = partner_code
                
            elif is_partner_in_collection:
                print(f"\n   📌 SCENARIO 2: RM is ALSO Partner, reports to different Partner")
                
                partner_entry = await db["Partner"].find_one({"PartnerEmpCode": partner_code})
                
                if not partner_entry:
                    raise HTTPException(
                        status_code=500,
                        detail=f"Partner {partner_code} not found in Partner collection"
                    )
                
                l1_approver_code = partner_code
                l1_approver_name = partner_entry.get("Partner_Name", partner_name)
                pending_queue_code = partner_code
                
            else:
                print(f"\n   📌 SCENARIO 3: Regular RM (not a Partner)")
                
                partner_entry = await db["Partner"].find_one({"PartnerEmpCode": partner_code})
                
                if not partner_entry:
                    raise HTTPException(
                        status_code=500,
                        detail=f"Partner {partner_code} not found in Partner collection"
                    )
                
                l1_approver_code = partner_code
                l1_approver_name = partner_entry.get("Partner_Name", partner_name)
                pending_queue_code = partner_code
            
            status_doc = {
                "employeeId": emp_code,
                "employee_name": employee_name,
                "month_range": month_range,
                "payroll_month": month_range,
                "total_amount": amount,
                "limit": ope_limit,
                "submission_date": datetime.utcnow(),
                "submitter_type": "Reporting_Manager",
                "approval_status": {
                    "total_levels": 2,
                    "current_level": "L1",
                    "overall_status": "pending",
                    "L1": {
                        "status": False,
                        "level_name": "Partner",
                        "approver_code": l1_approver_code,
                        "approver_name": l1_approver_name
                    },
                    "L2": {
                        "status": False,
                        "level_name": "HR",
                        "approver_code": "JHS729",
                        "approver_name": "HR Department"
                    }
                }
            }
            
            await db["Status"].update_one(
                {"employeeId": emp_code, "month_range": month_range},
                {"$set": status_doc},
                upsert=True
            )
            
            await db["Pending"].update_one(
                {"ReportingEmpCode": pending_queue_code},
                {
                    "$addToSet": {"EmployeesCodes": emp_code},
                    "$set": {"last_updated": datetime.utcnow()}
                },
                upsert=True
            )
            
            print(f"\n{'='*60}")
            print(f"✅ RM SUBMISSION COMPLETE")
            print(f"   Total Levels: 2 (ALWAYS)")
            print(f"   L1: Partner - {l1_approver_name} ({l1_approver_code})")
            print(f"   L2: HR (JHS729)")
            print(f"   Added to: {pending_queue_code}'s Pending")
            print(f"{'='*60}\n")
            
            return {
                "message": "RM entry submitted successfully",
                "entry_id": entry_id,
                "approval_flow": "2-level (Partner → HR)",
                "first_approver": f"Partner - {l1_approver_name}",
                "total_levels": 2,
                "submitter_type": "Reporting_Manager",
                "scenario": "RM_Self_Submission"
            }
        
        else:
            print(f"\n{'='*60}")
            print(f"👤 REGULAR EMPLOYEE SUBMISSION")
            print(f"{'='*60}")
            print(f"   Employee: {emp_code} - {employee_name}")
            print(f"   Reports to RM: {reporting_manager_code} - {reporting_manager_name}")
            print(f"   Partner: {partner_code} - {partner_name}")
            
            if amount > ope_limit:
                total_levels = 3
                
                print(f"\n   📌 SCENARIO 4: Employee Amount EXCEEDS Limit")
                print(f"      Amount: ₹{amount}")
                print(f"      Limit: ₹{ope_limit}")
                print(f"      ✅ Using 3-LEVEL approval")
                
                status_doc = {
                    "employeeId": emp_code,
                    "employee_name": employee_name,
                    "month_range": month_range,
                    "payroll_month": month_range,
                    "total_amount": amount,
                    "limit": ope_limit,
                    "submission_date": datetime.utcnow(),
                    "submitter_type": "Employee",
                    "approval_status": {
                        "total_levels": 3,
                        "current_level": "L1",
                        "overall_status": "pending",
                        "L1": {
                            "status": False,
                            "level_name": "Reporting Manager",
                            "approver_code": reporting_manager_code,
                            "approver_name": reporting_manager_name
                        },
                        "L2": {
                            "status": False,
                            "level_name": "Partner",
                            "approver_code": partner_code,
                            "approver_name": partner_name
                        },
                        "L3": {
                            "status": False,
                            "level_name": "HR",
                            "approver_code": "JHS729",
                            "approver_name": "HR Department"
                        }
                    }
                }
                
                await db["Pending"].update_one(
                    {"ReportingEmpCode": reporting_manager_code},
                    {
                        "$addToSet": {"EmployeesCodes": emp_code},
                        "$set": {"last_updated": datetime.utcnow()}
                    },
                    upsert=True
                )
                
                scenario_name = "Employee_Amount_Exceeds_Limit"
                
            else:
                total_levels = 2
                
                print(f"\n   📌 SCENARIO 5: Employee Amount WITHIN Limit")
                print(f"      Amount: ₹{amount}")
                print(f"      Limit: ₹{ope_limit}")
                print(f"      ✅ Using 2-LEVEL approval")
                
                status_doc = {
                    "employeeId": emp_code,
                    "employee_name": employee_name,
                    "month_range": month_range,
                    "payroll_month": month_range,
                    "total_amount": amount,
                    "limit": ope_limit,
                    "submission_date": datetime.utcnow(),
                    "submitter_type": "Employee",
                    "approval_status": {
                        "total_levels": 2,
                        "current_level": "L1",
                        "overall_status": "pending",
                        "L1": {
                            "status": False,
                            "level_name": "Reporting Manager",
                            "approver_code": reporting_manager_code,
                            "approver_name": reporting_manager_name
                        },
                        "L2": {
                            "status": False,
                            "level_name": "HR",
                            "approver_code": "JHS729",
                            "approver_name": "HR Department"
                        }
                    }
                }
                
                await db["Pending"].update_one(
                    {"ReportingEmpCode": reporting_manager_code},
                    {
                        "$addToSet": {"EmployeesCodes": emp_code},
                        "$set": {"last_updated": datetime.utcnow()}
                    },
                    upsert=True
                )
                
                scenario_name = "Employee_Amount_Within_Limit"
            
            await db["Status"].update_one(
                {"employeeId": emp_code, "month_range": month_range},
                {"$set": status_doc},
                upsert=True
            )
            
            print(f"\n{'='*60}")
            print(f"✅ EMPLOYEE SUBMISSION COMPLETE")
            print(f"   Scenario: {scenario_name}")
            print(f"   Total Levels: {total_levels}")
            print(f"   First Approver: RM - {reporting_manager_name}")
            print(f"{'='*60}\n")
            
            return {
                "message": "Entry submitted successfully",
                "entry_id": entry_id,
                "approval_flow": f"{total_levels}-level approval",
                "first_approver": f"RM - {reporting_manager_name}",
                "total_levels": total_levels,
                "submitter_type": "Employee",
                "scenario": scenario_name
            }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Submission error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

    
# ---------- GET HISTORY ----------
@app.get("/api/ope/history/{employee_code}")
async def get_ope_history(employee_code: str, current_user=Depends(get_current_user)):
    try:
        print(f"📌 Fetching history for: {employee_code}")
        
        if current_user["employee_code"] != employee_code:
            raise HTTPException(status_code=403, detail="Access denied")
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_code})
        
        if not ope_doc:
            return {"history": []}
        
        history = []
        data_array = ope_doc.get("Data", [])
        
        for data_item in data_array:
            for month_range, entries in data_item.items():
                for entry in entries:
                    history.append({
                        "_id": str(entry.get("_id", "")),
                        "month_range": month_range,
                        "date": entry.get("date"),
                        "client": entry.get("client"),
                        "project_id": entry.get("project_id"),
                        "project_name": entry.get("project_name"),
                        "project_type": entry.get("project_type", "N/A"), 
                        "location_from": entry.get("location_from"),
                        "location_to": entry.get("location_to"),
                        "travel_mode": entry.get("travel_mode"),
                        "amount": entry.get("amount"),
                        "remarks": entry.get("remarks"),
                        "ticket_pdf": entry.get("ticket_pdf"),   # GridFS ID or None
                        "created_time": entry.get("created_time"),
                        "updated_time": entry.get("updated_time")
                    })
        
        print(f"✅ Found {len(history)} entries")
        return {"history": history}
        
    except Exception as e:
        print(f"❌ Error fetching history: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ---------- UPDATE ENTRY ----------
@app.put("/api/ope/update/{entry_id}")
async def update_ope_entry(
    entry_id: str,
    update_data: dict,
    current_user=Depends(get_current_user)
):
    try:
        employee_code = current_user["employee_code"]
        print(f"📌 Updating entry {entry_id} for: {employee_code}")
        
        month_range = update_data.get("month_range")
        if not month_range:
            raise HTTPException(status_code=400, detail="month_range required")
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_code})
        if not ope_doc:
            raise HTTPException(status_code=404, detail="Employee data not found")
        
        data_array = ope_doc.get("Data", [])
        updated = False
        
        for i, data_item in enumerate(data_array):
            if month_range in data_item:
                entries = data_item[month_range]
                for j, entry in enumerate(entries):
                    if str(entry.get("_id")) == entry_id:
                        update_fields = {
                            f"Data.{i}.{month_range}.{j}.date": update_data.get("date"),
                            f"Data.{i}.{month_range}.{j}.client": update_data.get("client"),
                            f"Data.{i}.{month_range}.{j}.project_id": update_data.get("project_id"),
                            f"Data.{i}.{month_range}.{j}.project_name": update_data.get("project_name"),
                            f"Data.{i}.{month_range}.{j}.project_type": update_data.get("project_type"),
                            f"Data.{i}.{month_range}.{j}.location_from": update_data.get("location_from"),
                            f"Data.{i}.{month_range}.{j}.location_to": update_data.get("location_to"),
                            f"Data.{i}.{month_range}.{j}.travel_mode": update_data.get("travel_mode"),
                            f"Data.{i}.{month_range}.{j}.amount": update_data.get("amount"),
                            f"Data.{i}.{month_range}.{j}.remarks": update_data.get("remarks"),
                            f"Data.{i}.{month_range}.{j}.updated_time": datetime.utcnow().isoformat()
                        }
                        
                        await db["OPE_data"].update_one(
                            {"employeeId": employee_code},
                            {"$set": update_fields}
                        )
                        updated = True
                        break
            if updated:
                break
        
        if not updated:
            raise HTTPException(status_code=404, detail="Entry not found")
        
        print(f"✅ Entry updated successfully")
        return {"message": "Entry updated successfully"}
        
    except Exception as e:
        print(f"❌ Error updating entry: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ---------- DELETE ENTRY ----------
@app.delete("/api/ope/delete/{entry_id}")
async def delete_ope_entry(
    entry_id: str,
    delete_data: dict,
    current_user=Depends(get_current_user)
):
    try:
        employee_code = current_user["employee_code"]
        print(f"📌 Deleting entry {entry_id} for: {employee_code}")
        
        if not entry_id or entry_id == "dummy":
            raise HTTPException(status_code=400, detail="Invalid entry ID")
        
        month_range = delete_data.get("month_range")
        if not month_range:
            raise HTTPException(status_code=400, detail="month_range required")
        
        print(f"📌 Month range: {month_range}")
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_code})
        
        if not ope_doc:
            raise HTTPException(status_code=404, detail="Employee data not found")
        
        data_array = ope_doc.get("Data", [])
        deleted = False
        
        for i, data_item in enumerate(data_array):
            if month_range in data_item:
                entries = data_item[month_range]
                
                for j, entry in enumerate(entries):
                    if str(entry.get("_id")) == entry_id:
                        print(f"✅ Found entry at Data.{i}.{month_range}.{j}")
                        
                        # Delete PDF from GridFS if exists
                        pdf_id = entry.get("ticket_pdf")
                        if pdf_id and is_gridfs_id(pdf_id):
                            await delete_from_gridfs(pdf_id)
                            print(f"✅ PDF deleted from GridFS: {pdf_id}")
                        
                        if len(entries) == 1:
                            print(f"🗑️ Removing entire month range: {month_range}")
                            await db["OPE_data"].update_one(
                                {"employeeId": employee_code},
                                {"$pull": {"Data": {month_range: {"$exists": True}}}}
                            )
                        else:
                            print(f"🗑️ Removing single entry from month range")
                            await db["OPE_data"].update_one(
                                {"employeeId": employee_code},
                                {"$pull": {f"Data.{i}.{month_range}": {"_id": ObjectId(entry_id)}}}
                            )
                        
                        deleted = True
                        break
            
            if deleted:
                break
        
        if not deleted:
            raise HTTPException(status_code=404, detail="Entry not found")
        
        print(f"✅ Entry deleted successfully")
        return {
            "message": "Entry deleted successfully",
            "entry_id": entry_id,
            "month_range": month_range
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error deleting entry: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    
@app.get("/api/check-role/{employee_code}")
async def check_user_role(employee_code: str, current_user=Depends(get_current_user)):
    try:
        emp_code = employee_code.strip().upper()
        
        if current_user["employee_code"].upper() != emp_code:
            raise HTTPException(status_code=403, detail="Access denied")
        
        is_hr = (emp_code == "JHS729")
        
        partner = await db["Partner"].find_one({"PartnerEmpCode": emp_code})
        is_partner = partner is not None
        
        manager = await db["Reporting_managers"].find_one({"ReportingEmpCode": emp_code})
        is_manager = manager is not None
        
        return {
            "employee_code": emp_code,
            "is_hr": is_hr,
            "is_partner": is_partner,
            "is_manager": is_manager,
            "is_employee": not (is_hr or is_partner or is_manager)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    

@app.get("/api/check-manager/{employee_code}")
async def check_if_manager(employee_code: str, current_user=Depends(get_current_user)):
    try:
        emp_code = employee_code.strip().upper()
        
        print(f"🔍 Checking if {emp_code} is a manager...")
        
        if current_user["employee_code"].upper() != emp_code:
            raise HTTPException(status_code=403, detail="Access denied")
        
        manager = await db["Reporting_managers"].find_one({"ReportingEmpCode": emp_code})
        
        is_manager = manager is not None
        
        if is_manager:
            print(f"✅ {emp_code} IS a reporting manager")
        else:
            print(f"❌ {emp_code} is NOT a reporting manager")
        
        return {
            "employee_code": emp_code,
            "isManager": is_manager,
            "manager_name": manager.get("ReportingEmpName") if manager else None,
            "email": manager.get("Email ID") if manager else None
        }
        
    except Exception as e:
        print(f"❌ Error checking manager role: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/ope/manager/employees/{status}")
async def get_employees_by_status(
    status: str,
    current_user=Depends(get_current_user)
):
    try:
        reporting_emp_code = current_user["employee_code"].strip().upper()
        
        print(f"🔍 Fetching {status} employees for manager: {reporting_emp_code}")
        
        manager = await db["Reporting_managers"].find_one({"ReportingEmpCode": reporting_emp_code})
        if not manager:
            raise HTTPException(status_code=403, detail="You are not a reporting manager")
        
        status_lower = status.lower()
        if status_lower == "pending":
            collection_name = "Pending"
        elif status_lower == "approved":
            collection_name = "Approved"
        elif status_lower == "rejected":
            collection_name = "Rejected"
        else:
            raise HTTPException(status_code=400, detail="Invalid status. Use: pending, approved, or rejected")
        
        status_collection = db[collection_name]
        status_doc = await status_collection.find_one({"ReportingEmpCode": reporting_emp_code})
        
        print(f"📄 Status doc found: {status_doc is not None}")
        
        if not status_doc:
            return {
                "message": f"No {status} data found for manager: {reporting_emp_code}",
                "reporting_manager": reporting_emp_code,
                "employees": []
            }
        
        employee_codes = status_doc.get("EmployeesCodes", [])
        print(f"👥 Employee codes: {employee_codes}")
        
        employees_data = []
        
        for emp_code in employee_codes:
            ope_data = await db["OPE_data"].find_one(
                {"employeeId": emp_code},
                {"_id": 0}
            )
            
            print(f"📊 OPE data for {emp_code}: {ope_data is not None}")
            
            if ope_data:
                employees_data.append({
                    "employeeId": emp_code,
                    "employeeName": ope_data.get("employeeName", ""),
                    "designation": ope_data.get("designation", ""),
                    "opeData": ope_data
                })
            else:
                employees_data.append({
                    "employeeId": emp_code,
                    "employeeName": "Unknown",
                    "designation": "Unknown",
                    "opeData": None
                })
        
        print(f"✅ Returning {len(employees_data)} employees")
        
        return {
            "reporting_manager": reporting_emp_code,
            "status": status_lower,
            "total_employees": len(employees_data),
            "employees": employees_data
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error fetching employees: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/ope/manager/pending")
async def get_manager_pending_employees(current_user=Depends(get_current_user)):
    try:
        current_emp_code = current_user["employee_code"].strip().upper()
        
        print(f"\n{'='*60}")
        print(f"🔍 PENDING REQUEST FROM: {current_emp_code}")
        print(f"{'='*60}\n")
        
        is_hr = (current_emp_code == "JHS729")
        
        if is_hr:
            print(f"👔 USER IS HR - Fetching L1/L2 approved entries")
            
            all_status_docs = await db["Status"].find({}).to_list(length=None)
            print(f"📊 Total Status documents in DB: {len(all_status_docs)}")
            
            pending_employees = []
            
            for status_doc in all_status_docs:
                employee_id = status_doc.get("employeeId")
                employee_name = status_doc.get("employeeName", "Unknown")
                approval_status = status_doc.get("approval_status", [])
                
                print(f"\n📋 Checking Employee: {employee_id} ({employee_name})")
                print(f"   Total payroll months: {len(approval_status)}")
                
                ope_doc = await db["OPE_data"].find_one({"employeeId": employee_id})
                if not ope_doc:
                    print(f"   ⚠️ No OPE_data found - skipping")
                    continue
                
                pending_entries = []
                
                for ps_index, ps in enumerate(approval_status):
                    payroll_month = ps.get("payroll_month")
                    total_levels = ps.get("total_levels", 2)
                    current_level = ps.get("current_level", "L1")
                    overall_status = ps.get("overall_status", "pending")
                    L1 = ps.get("L1", {})
                    L2 = ps.get("L2", {})
                    
                    print(f"\n   📅 Payroll: {payroll_month}")
                    print(f"      Total Levels: {total_levels}")
                    print(f"      Current Level: {current_level}")
                    print(f"      Overall Status: {overall_status}")
                    print(f"      L1 Status: {L1.get('status')}")
                    print(f"      L2 Status: {L2.get('status')}")
                    
                    should_show_to_hr = False
                    
                    if total_levels == 2:
                        if L1.get("status") == True and current_level == "L2" and overall_status == "pending":
                            should_show_to_hr = True
                            print(f"      ✅ MATCH: 2-level pending at HR (L1 approved)")
                    
                    elif total_levels == 3:
                        L3 = ps.get("L3", {})
                        print(f"      L3 Status: {L3.get('status')}")
                        if (L1.get("status") == True and 
                            L2.get("status") == True and 
                            current_level == "L3" and 
                            overall_status == "pending"):
                            should_show_to_hr = True
                            print(f"      ✅ MATCH: 3-level pending at HR (L1+L2 approved)")
                    
                    if not should_show_to_hr:
                        print(f"      ❌ NOT for HR - skipping")
                        continue
                    
                    data_array = ope_doc.get("Data", [])
                    for data_item in data_array:
                        if payroll_month in data_item:
                            entries = data_item[payroll_month]
                            print(f"      📦 Found {len(entries)} entries in OPE_data")
                            for entry in entries:
                                entry_status = entry.get("status", "").lower()
                                if entry_status == "approved":
                                    pending_entries.append({
                                        "_id": str(entry.get("_id", "")),
                                        "month_range": payroll_month,
                                        "date": entry.get("date"),
                                        "client": entry.get("client"),
                                        "project_id": entry.get("project_id"),
                                        "project_name": entry.get("project_name"),
                                        "project_type": entry.get("project_type", "N/A"),
                                        "location_from": entry.get("location_from"),
                                        "location_to": entry.get("location_to"),
                                        "travel_mode": entry.get("travel_mode"),
                                        "amount": entry.get("amount"),
                                        "remarks": entry.get("remarks"),
                                        "ticket_pdf": entry.get("ticket_pdf"),    # GridFS ID
                                        "total_levels": total_levels,
                                        "current_level": current_level
                                    })
                                    print(f"         ✅ Entry added: {entry.get('date')} - ₹{entry.get('amount')}")
                                else:
                                    print(f"         ⚠️ Entry skipped - status: {entry_status}")
                            break
                
                if pending_entries:
                    pending_employees.append({
                        "employeeId": employee_id,
                        "employeeName": employee_name,
                        "designation": ope_doc.get("designation", ""),
                        "pendingCount": len(pending_entries),
                        "entries": pending_entries
                    })
                    print(f"\n   ✅ ADDED: {employee_name} with {len(pending_entries)} pending entries")
                else:
                    print(f"   ❌ No pending entries for HR")
            
            print(f"\n{'='*60}")
            print(f"✅ FINAL RESULT: {len(pending_employees)} employees pending for HR")
            print(f"{'='*60}\n")
            
            return {
                "reporting_manager": current_emp_code,
                "is_hr": True,
                "total_employees": len(pending_employees),
                "employees": pending_employees
            }

        else:
            # ✅ CHECK PARTNER FIRST
            is_partner = await db["Partner"].find_one({"PartnerEmpCode": current_emp_code})

            if is_partner:
                print(f"🤝 USER IS PARTNER")

                partner_pending = await db["Pending"].find_one({"ReportingEmpCode": current_emp_code})
                if not partner_pending or not partner_pending.get("EmployeesCodes"):
                    print(f"✅ No pending employees found for partner")
                    return {"reporting_manager": current_emp_code, "is_hr": False, "is_partner": True, "total_employees": 0, "employees": []}

                pending_emp_codes = partner_pending.get("EmployeesCodes", [])
                print(f"👥 Found {len(pending_emp_codes)} employees under partner")
                pending_employees = []

                for emp_code in pending_emp_codes:
                    status_docs = await db["Status"].find({"employeeId": emp_code}).to_list(length=None)
                    if not status_docs:
                        continue

                    ope_doc = await db["OPE_data"].find_one({"employeeId": emp_code})
                    if not ope_doc:
                        continue

                    pending_entries = []

                    for status_doc in status_docs:
                        approval_status_raw = status_doc.get("approval_status")
                        if isinstance(approval_status_raw, list):
                            approval_status_array = approval_status_raw
                        elif isinstance(approval_status_raw, dict):
                            approval_status_array = [approval_status_raw]
                        else:
                            continue

                        for approval_status in approval_status_array:
                            month_range = approval_status.get("payroll_month") or approval_status.get("month_range")
                            total_levels = approval_status.get("total_levels", 2)
                            current_level = approval_status.get("current_level", "L1")
                            overall_status = approval_status.get("overall_status", "pending")
                            submitter_type = approval_status.get("submitter_type", "Employee")

                            if overall_status != "pending" or not month_range:
                                continue

                            partner_is_approver = False
                            if current_level == "L1":
                                L1 = approval_status.get("L1", {})
                                if L1.get("approver_code") == current_emp_code:
                                    partner_is_approver = True
                            elif current_level == "L2":
                                L2 = approval_status.get("L2", {})
                                if L2.get("approver_code") == current_emp_code:
                                    partner_is_approver = True

                            if not partner_is_approver:
                                continue

                            print(f"   ✅ Partner is approver for {emp_code} - {month_range} (Level: {current_level})")

                            for data_item in ope_doc.get("Data", []):
                                if month_range in data_item:
                                    target_status = "pending" if submitter_type == "Reporting_Manager" else "approved"
                                    for entry in data_item[month_range]:
                                        e_status = entry.get("status", "").lower()
                                        if e_status in [target_status, "approved", "pending"]:
                                            pending_entries.append({
                                                "_id": str(entry.get("_id", "")),
                                                "month_range": month_range,
                                                "date": entry.get("date"),
                                                "client": entry.get("client"),
                                                "project_id": entry.get("project_id"),
                                                "project_name": entry.get("project_name"),
                                                "project_type": entry.get("project_type", "N/A"),
                                                "location_from": entry.get("location_from"),
                                                "location_to": entry.get("location_to"),
                                                "travel_mode": entry.get("travel_mode"),
                                                "amount": entry.get("amount"),
                                                "remarks": entry.get("remarks"),
                                                "ticket_pdf": entry.get("ticket_pdf"),   # GridFS ID
                                                "total_levels": total_levels,
                                                "current_level": current_level
                                            })
                                    break

                    if pending_entries:
                        emp_info = await db["Employee_details"].find_one({"EmpID": emp_code}) or {}
                        pending_employees.append({
                            "employeeId": emp_code,
                            "employeeName": emp_info.get("Emp Name", ope_doc.get("employeeName", emp_code)),
                            "designation": emp_info.get("Designation Name", ope_doc.get("designation", "")),
                            "pendingCount": len(pending_entries),
                            "entries": pending_entries
                        })

                print(f"✅ Returning {len(pending_employees)} employees for partner")
                return {
                    "reporting_manager": current_emp_code,
                    "is_hr": False,
                    "is_partner": True,
                    "total_employees": len(pending_employees),
                    "employees": pending_employees
                }

            # ✅ REPORTING MANAGER LOGIC
            print(f"👔 USER IS REPORTING MANAGER")

            manager = await db["Reporting_managers"].find_one({"ReportingEmpCode": current_emp_code})
            if not manager:
                raise HTTPException(status_code=403, detail="You are not a reporting manager")

            employees = await db["Employee_details"].find(
                {"ReportingEmpCode": current_emp_code}
            ).to_list(length=None)

            print(f"👥 Found {len(employees)} employees under manager")

            pending_employees = []

            for emp in employees:
                emp_code = emp.get("EmpID")
                emp_name = emp.get("Emp Name")

                ope_doc = await db["OPE_data"].find_one({"employeeId": emp_code})
                if ope_doc:
                    pending_entries = []
                    data_array = ope_doc.get("Data", [])

                    for data_item in data_array:
                        for month_range, entries in data_item.items():
                            for entry in entries:
                                entry_status = entry.get("status", "pending").lower()
                                if entry_status == "pending":
                                    pending_entries.append({
                                        "_id": str(entry.get("_id", "")),
                                        "month_range": month_range,
                                        "date": entry.get("date"),
                                        "client": entry.get("client"),
                                        "project_id": entry.get("project_id"),
                                        "project_name": entry.get("project_name"),
                                        "project_type": entry.get("project_type", "N/A"),
                                        "location_from": entry.get("location_from"),
                                        "location_to": entry.get("location_to"),
                                        "travel_mode": entry.get("travel_mode"),
                                        "amount": entry.get("amount"),
                                        "remarks": entry.get("remarks"),
                                        "ticket_pdf": entry.get("ticket_pdf")    # GridFS ID
                                    })

                    if pending_entries:
                        pending_employees.append({
                            "employeeId": emp_code,
                            "employeeName": emp_name,
                            "designation": emp.get("Designation Name", ""),
                            "pendingCount": len(pending_entries),
                            "entries": pending_entries
                        })

            print(f"✅ Returning {len(pending_employees)} employees for manager")

            return {
                "reporting_manager": current_emp_code,
                "is_hr": False,
                "total_employees": len(pending_employees),
                "employees": pending_employees
            }

    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    

@app.get("/api/ope/approved/{employee_code}")
async def get_employee_approved(
    employee_code: str, 
    current_user=Depends(get_current_user)
):
    try:
        employee_code = employee_code.strip().upper()
        current_emp_code = current_user["employee_code"].strip().upper()
        
        print(f"\n{'='*60}")
        print(f"📊 GET APPROVED ENTRIES")
        print(f"   Employee: {employee_code}")
        print(f"   Current user: {current_emp_code}")
        print(f"{'='*60}\n")
        
        is_hr = (current_emp_code == "JHS729")
        is_own_data = (current_emp_code == employee_code)
        is_manager = await db["Reporting_managers"].find_one(
            {"ReportingEmpCode": current_emp_code}
        )
        
        if not (is_hr or is_own_data or is_manager):
            print(f"❌ Access denied - Not HR, not own data, and not a manager")
            raise HTTPException(status_code=403, detail="Access denied")
        
        print(f"✅ Access granted - Fetching OPE data")
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_code})
        
        if not ope_doc:
            print(f"📭 No OPE data found for {employee_code}")
            return {"approved": []}
        
        print(f"✅ OPE document found")
        
        approved_entries = []
        data_array = ope_doc.get("Data", [])
        
        print(f"📊 Total data items: {len(data_array)}")
        
        for data_item in data_array:
            for month_range, entries in data_item.items():
                print(f"   📅 Month: {month_range}, Entries: {len(entries)}")
                
                for entry in entries:
                    entry_status = entry.get("status", "").lower()
                    print(f"      Entry ID: {entry.get('_id')}, Status: '{entry_status}'")
                    
                    if entry_status == "approved":
                        approval_remark = entry.get("approval_remark") or entry.get("remark") or ""
                        
                        approved_entries.append({
                            "_id": str(entry.get("_id", "")),
                            "employee_id": employee_code,
                            "employee_name": ope_doc.get("employeeName", ""),
                            "designation": ope_doc.get("designation", ""),
                            "month_range": month_range,
                            "date": entry.get("date"),
                            "client": entry.get("client"),
                            "project_id": entry.get("project_id"),
                            "project_name": entry.get("project_name"),
                            "project_type": entry.get("project_type", "N/A"),
                            "location_from": entry.get("location_from"),
                            "location_to": entry.get("location_to"),
                            "travel_mode": entry.get("travel_mode"),
                            "amount": entry.get("amount"),
                            "remarks": entry.get("remarks"),
                            "ticket_pdf": entry.get("ticket_pdf"),    # GridFS ID
                            "approved_by": entry.get("approved_by"),
                            "approver_name": entry.get("approver_name"),
                            "approval_remark": approval_remark,
                            "approved_date": entry.get("approved_date"),
                            "created_time": entry.get("created_time"),
                            "L1_approved": entry.get("L1_approved"),
                            "L1_approver_code": entry.get("L1_approver_code"),
                            "L1_approver_name": entry.get("L1_approver_name")
                        })
                        print(f"      ✅ APPROVED entry added with remark: {approval_remark}")
        
        print(f"\n✅ Total approved entries found: {len(approved_entries)}\n")
        
        return {"approved": approved_entries}
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"\n❌❌ ERROR fetching approved:")
        print(f"   {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/ope/rejected/{employee_code}")
async def get_employee_rejected(
    employee_code: str, 
    current_user=Depends(get_current_user)
):
    try:
        employee_code = employee_code.strip().upper()
        current_emp_code = current_user["employee_code"].strip().upper()
        
        print(f"\n{'='*60}")
        print(f"❌ GET REJECTED ENTRIES")
        print(f"   Employee: {employee_code}")
        print(f"   Current user: {current_emp_code}")
        print(f"{'='*60}\n")
        
        is_hr = (current_emp_code == "JHS729")
        is_own_data = (current_emp_code == employee_code)
        is_manager = await db["Reporting_managers"].find_one(
            {"ReportingEmpCode": current_emp_code}
        )
        
        if not (is_hr or is_own_data or is_manager):
            print(f"❌ Access denied")
            raise HTTPException(status_code=403, detail="Access denied")
        
        print(f"✅ Access granted - Fetching OPE data")
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_code})
        
        if not ope_doc:
            print(f"📭 No OPE data found for {employee_code}")
            return {"rejected": []}
        
        print(f"✅ OPE document found")
        
        rejected_entries = []
        data_array = ope_doc.get("Data", [])
        
        print(f"📊 Total data items: {len(data_array)}")
        
        for data_item in data_array:
            for month_range, entries in data_item.items():
                print(f"   📅 Month: {month_range}, Entries: {len(entries)}")
                
                for entry in entries:
                    entry_status = entry.get("status", "").lower()
                    print(f"      Entry ID: {entry.get('_id')}, Status: '{entry_status}'")
                    
                    if entry_status == "rejected":
                        rejected_entries.append({
                            "_id": str(entry.get("_id", "")),
                            "employee_id": employee_code,
                            "employee_name": ope_doc.get("employeeName", ""),
                            "designation": ope_doc.get("designation", ""),
                            "month_range": month_range,
                            "date": entry.get("date"),
                            "client": entry.get("client"),
                            "project_id": entry.get("project_id"),
                            "project_name": entry.get("project_name"),
                            "project_type": entry.get("project_type", "N/A"),
                            "location_from": entry.get("location_from"),
                            "location_to": entry.get("location_to"),
                            "travel_mode": entry.get("travel_mode"),
                            "amount": entry.get("amount"),
                            "remarks": entry.get("remarks"),
                            "ticket_pdf": entry.get("ticket_pdf"),    # GridFS ID
                            "rejected_by": entry.get("rejected_by"),
                            "rejector_name": entry.get("rejector_name"),
                            "rejected_date": entry.get("rejected_date"),
                            "rejection_reason": entry.get("rejection_reason"),
                            "rejected_level": entry.get("rejected_level"),
                            "created_time": entry.get("created_time")
                        })
                        print(f"      ✅ REJECTED entry added")
        
        print(f"\n✅ Total rejected entries found: {len(rejected_entries)}\n")
        return {"rejected": rejected_entries}
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"\n❌❌ ERROR fetching rejected:")
        print(f"   {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))



@app.post("/api/ope/manager/reject/{employee_code}")
async def reject_employee_entries(
    employee_code: str,
    request: Request,
    current_user=Depends(get_current_user)
):
    try:
        reporting_emp_code = current_user["employee_code"].strip().upper()
        employee_code = employee_code.strip().upper()
        
        body = await request.json()
        rejection_reason = body.get("reason", "No reason provided")
        
        print(f"\n{'='*60}")
        print(f"❌ REJECTING EMPLOYEE")
        print(f"Manager: {reporting_emp_code}")
        print(f"Employee: {employee_code}")
        print(f"Reason: {rejection_reason}")
        print(f"{'='*60}\n")
        
        manager = await db["Reporting_managers"].find_one({"ReportingEmpCode": reporting_emp_code})
        if not manager:
            raise HTTPException(status_code=403, detail="You are not a reporting manager")
        
        manager_name = manager.get("ReportingEmpName", reporting_emp_code)
        
        emp = await db["Employee_details"].find_one({"EmpID": employee_code})
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        
        emp_reporting_manager_code = emp.get("ReportingEmpCode", "").strip().upper()
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_code})
        if not ope_doc:
            raise HTTPException(status_code=404, detail="No data found")
        
        data_array = ope_doc.get("Data", [])
        rejected_count = 0
        current_time = datetime.utcnow().isoformat()
        
        for i, data_item in enumerate(data_array):
            for month_range, entries in data_item.items():
                for j, entry in enumerate(entries):
                    if entry.get("status", "").lower() == "pending":
                        await db["OPE_data"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"Data.{i}.{month_range}.{j}.status": "rejected",
                                f"Data.{i}.{month_range}.{j}.rejected_by": reporting_emp_code,
                                f"Data.{i}.{month_range}.{j}.rejector_name": manager_name,
                                f"Data.{i}.{month_range}.{j}.rejected_date": current_time,
                                f"Data.{i}.{month_range}.{j}.rejection_reason": rejection_reason
                            }}
                        )
                        
                        status_id = entry.get("status_id")
                        if status_id:
                            await db["Status"].update_one(
                                {"_id": ObjectId(status_id)},
                                {"$set": {
                                    "overall_status": "rejected",
                                    "L1.status": False,
                                    "L1.rejected_by": reporting_emp_code,
                                    "L1.rejected_date": current_time
                                }}
                            )
                        
                        rejected_count += 1
                        print(f"✅ Rejected entry {j + 1}")
        
        if rejected_count == 0:
            raise HTTPException(status_code=404, detail="No pending entries found")
        
        print(f"\n✅ Total entries rejected: {rejected_count}")
        
        await db["Pending"].update_one(
            {"ReportingEmpCode": emp_reporting_manager_code},
            {"$pull": {"EmployeesCodes": employee_code}}
        )
        print(f"✅ Removed from Pending collection")
        
        rejected_doc = await db["Rejected"].find_one({"ReportingEmpCode": reporting_emp_code})
        
        if not rejected_doc:
            await db["Rejected"].insert_one({
                "ReportingEmpCode": reporting_emp_code,
                "EmployeesCodes": [employee_code]
            })
            print(f"✅ Created NEW Rejected document")
        else:
            if employee_code not in rejected_doc.get("EmployeesCodes", []):
                await db["Rejected"].update_one(
                    {"ReportingEmpCode": reporting_emp_code},
                    {"$addToSet": {"EmployeesCodes": employee_code}}
                )
                print(f"✅ Added to Rejected collection")
            else:
                print(f"⚠️ Employee already in Rejected collection")
        
        print(f"{'='*60}\n")
        
        return {
            "message": f"Rejected {rejected_count} entries",
            "rejected_count": rejected_count,
            "rejection_reason": rejection_reason,
            "employee_code": employee_code
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))  


@app.put("/api/ope/manager/edit-total-amount")
async def edit_total_amount(
    request: Request,
    current_user=Depends(get_current_user)
):
    try:
        user_emp_code = current_user["employee_code"].strip().upper()
        
        body = await request.json()
        employee_id = body.get("employee_id")
        month_range = body.get("month_range")
        new_total = body.get("new_total")
        
        print(f"\n{'='*60}")
        print(f"💰💰 EDIT TOTAL AMOUNT REQUEST")
        print(f"   User: {user_emp_code}")
        print(f"   Employee: {employee_id}")
        print(f"   Month: {month_range}")
        print(f"   New Total: ₹{new_total}")
        print(f"{'='*60}\n")
        
        is_manager = await db["Reporting_managers"].find_one({"ReportingEmpCode": user_emp_code})
        is_hr = (user_emp_code == "JHS729")
        is_partner = await db["Partner"].find_one({"PartnerEmpCode": user_emp_code})
        
        if not (is_manager or is_hr or is_partner):
            raise HTTPException(status_code=403, detail="Only managers, partners, and HR can edit amounts")
        
        if is_hr:
            user_role = "HR"
            user_name = "HR Department"
        elif is_partner:
            partner = await db["Partner"].find_one({"PartnerEmpCode": user_emp_code})
            user_role = "Partner"
            user_name = partner.get("Partner_Name", user_emp_code) if partner else user_emp_code
        else:
            manager = await db["Reporting_managers"].find_one({"ReportingEmpCode": user_emp_code})
            user_role = "Manager"
            user_name = manager.get("ReportingEmpName", user_emp_code) if manager else user_emp_code
        
        if not employee_id or not month_range or new_total is None:
            raise HTTPException(status_code=400, detail="Missing required fields")
        
        if new_total <= 0:
            raise HTTPException(status_code=400, detail="Total amount must be greater than 0")
        
        current_time = datetime.utcnow().isoformat()
        
        status_doc = await db["Status"].find_one({"employeeId": employee_id})
        
        if not status_doc:
            raise HTTPException(status_code=404, detail="Status document not found")
        
        approval_status = status_doc.get("approval_status", [])
        found = False
        old_total = 0
        
        for i, ps in enumerate(approval_status):
            if ps.get("payroll_month") == month_range:
                old_total = ps.get("total_amount", 0)
                
                edit_history = ps.get("amount_edit_history", [])
                edit_history.append({
                    "edited_by": user_emp_code,
                    "edited_by_name": user_name,
                    "edited_by_role": user_role,
                    "edited_date": current_time,
                    "old_total": old_total,
                    "new_total": new_total,
                    "entries_updated": 0
                })
                
                await db["Status"].update_one(
                    {"employeeId": employee_id},
                    {"$set": {
                        f"approval_status.{i}.total_amount": new_total,
                        f"approval_status.{i}.original_total": old_total,
                        f"approval_status.{i}.last_edited_by": user_emp_code,
                        f"approval_status.{i}.last_edited_by_name": user_name,
                        f"approval_status.{i}.last_edited_by_role": user_role,
                        f"approval_status.{i}.last_edited_date": current_time,
                        f"approval_status.{i}.amount_edit_history": edit_history
                    }}
                )
                found = True
                print(f"✅ Updated Status total: ₹{old_total} → ₹{new_total}")
                break
        
        if not found:
            raise HTTPException(status_code=404, detail="Payroll month not found in Status")
        
        return {
            "success": True,
            "message": "Total amount updated successfully. Individual entry amounts unchanged.",
            "updated_by": {
                "code": user_emp_code,
                "name": user_name,
                "role": user_role
            },
            "old_total": old_total,
            "new_total": new_total,
            "entries_updated": 0,
            "note": "Only the payroll total was updated. Entry-level amounts remain as submitted."
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error editing total amount: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))    


@app.put("/api/ope/manager/edit-amount")
async def edit_entry_amount(
    request: Request,
    current_user=Depends(get_current_user)
):
    try:
        user_emp_code = current_user["employee_code"].strip().upper()
        
        body = await request.json()
        entry_id = body.get("entry_id")
        employee_id = body.get("employee_id")
        new_amount = body.get("new_amount")
        
        print(f"\n{'='*60}")
        print(f"💰 EDIT SINGLE AMOUNT REQUEST")
        print(f"   User: {user_emp_code}")
        print(f"   Employee: {employee_id}")
        print(f"   Entry ID: {entry_id}")
        print(f"   New Amount: {new_amount}")
        print(f"{'='*60}\n")
        
        is_manager = await db["Reporting_managers"].find_one({"ReportingEmpCode": user_emp_code})
        is_hr = (user_emp_code == "JHS729")
        
        if not is_manager and not is_hr:
            error_msg = "Only managers and HR can edit amounts"
            print(f"❌ Authorization failed: {error_msg}")
            raise HTTPException(
                status_code=403,
                detail=error_msg
            )
        
        user_role = "HR" if is_hr else "Manager"
        print(f"✅ User role: {user_role}")
        
        if not entry_id or not employee_id or new_amount is None:
            raise HTTPException(status_code=400, detail="Missing required fields")
        
        if new_amount <= 0:
            raise HTTPException(status_code=400, detail="Amount must be greater than 0")
        
        emp = await db["Employee_details"].find_one({"EmpID": employee_id})
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_id})
        
        if not ope_doc:
            raise HTTPException(status_code=404, detail="Employee data not found")
        
        data_array = ope_doc.get("Data", [])
        updated = False
        old_amount = 0
        payroll_month = None
        current_time = datetime.utcnow().isoformat()
        
        for i, data_item in enumerate(data_array):
            for month_range, entries in data_item.items():
                for j, entry in enumerate(entries):
                    if str(entry.get("_id")) == entry_id:
                        old_amount = entry.get("amount", 0)
                        payroll_month = month_range
                        
                        await db["OPE_data"].update_one(
                            {"employeeId": employee_id},
                            {"$set": {
                                f"Data.{i}.{month_range}.{j}.amount": new_amount,
                                f"Data.{i}.{month_range}.{j}.updated_time": current_time,
                                f"Data.{i}.{month_range}.{j}.amount_edited_by": user_emp_code,
                                f"Data.{i}.{month_range}.{j}.amount_edited_by_role": user_role,
                                f"Data.{i}.{month_range}.{j}.amount_edited_date": current_time
                            }}
                        )
                        
                        updated = True
                        print(f"✅ Amount updated: ₹{old_amount} → ₹{new_amount}")
                        break
            if updated:
                break
        
        if not updated:
            raise HTTPException(status_code=404, detail="Entry not found")
        
        if payroll_month:
            status_doc = await db["Status"].find_one({"employeeId": employee_id})
            
            if status_doc:
                approval_status = status_doc.get("approval_status", [])
                
                new_total = 0
                for i, data_item in enumerate(data_array):
                    if payroll_month in data_item:
                        entries = data_item[payroll_month]
                        new_total = sum(float(e.get("amount", 0)) for e in entries)
                        break
                
                for i, ps in enumerate(approval_status):
                    if ps.get("payroll_month") == payroll_month:
                        await db["Status"].update_one(
                            {"employeeId": employee_id},
                            {"$set": {
                                f"approval_status.{i}.total_amount": new_total
                            }}
                        )
                        print(f"✅ Updated Status total_amount to: ₹{new_total}")
                        break
        
        print(f"{'='*60}\n")
        
        return {
            "success": True,
            "message": "Amount updated successfully",
            "updated_by_role": user_role,
            "old_amount": old_amount,
            "new_amount": new_amount,
            "entry_id": entry_id
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error editing amount: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ==================================================
# TEMPORARY SAVE ENDPOINT (Store in Temp_OPE_data)
# ==================================================

@app.post("/api/ope/save-temp")
async def save_temp_entry(
    date: str = Form(...),
    client: str = Form(...),
    project_id: str = Form(...),
    project_name: str = Form(...),
    project_type: str = Form(...),
    location_from: str = Form(...),
    location_to: str = Form(...), 
    travel_mode: str = Form(...),
    amount: float = Form(...),
    remarks: str = Form(...),
    month_range: str = Form(...),
    ticket_pdf: Optional[UploadFile] = File(None),
    current_user=Depends(get_current_user)
):
    try:
        employee_code = current_user["employee_code"]
        
        print(f"\n{'='*60}")
        print(f"💾 SAVING TEMPORARY ENTRY")
        print(f"Employee: {employee_code}")
        print(f"Amount: ₹{amount}")
        print(f"{'='*60}\n")
        
        emp = await db["Employee_details"].find_one({"EmpID": employee_code})
        
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        
        actual_ope_limit = emp.get("OPE LIMIT")
        
        if actual_ope_limit is None:
            print(f"⚠️ OPE Limit not found in database, using default")
            actual_ope_limit = 5000
        else:
            actual_ope_limit = float(actual_ope_limit)
        
        print(f"✅ Employee's Actual OPE Limit: ₹{actual_ope_limit}")
        
        def format_month_range(month_str):
            try:
                parts = month_str.lower().split('-')
                month_map = {
                    'jan': 'Jan', 'feb': 'Feb', 'mar': 'Mar', 'apr': 'Apr',
                    'may': 'May', 'jun': 'Jun', 'jul': 'Jul', 'aug': 'Aug',
                    'sep': 'Sep', 'oct': 'Oct', 'nov': 'Nov', 'dec': 'Dec'
                }
                
                if len(parts) == 3:
                    month1 = month_map.get(parts[0], parts[0].capitalize())
                    month2 = month_map.get(parts[1], parts[1].capitalize())
                    year = parts[2]
                    return f"{month1} {year} - {month2} {year}"
                elif len(parts) == 2:
                    month = month_map.get(parts[0], parts[0].capitalize())
                    year = parts[1]
                    return f"{month} {year}"
                else:
                    return month_str
            except Exception as e:
                return month_str
        
        formatted_month_range = format_month_range(month_range)
        
        # Get existing temp entries for this month
        temp_doc = await db["Temp_OPE_data"].find_one({"employeeId": employee_code})
        
        existing_month_total = 0
        
        if temp_doc:
            data_array = temp_doc.get("Data", [])
            for data_item in data_array:
                if formatted_month_range in data_item:
                    entries = data_item[formatted_month_range]
                    existing_month_total = sum(float(e.get("amount", 0)) for e in entries)
                    break
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_code})
        
        submitted_month_total = 0
        
        if ope_doc:
            data_array = ope_doc.get("Data", [])
            for data_item in data_array:
                if formatted_month_range in data_item:
                    entries = data_item[formatted_month_range]
                    submitted_month_total = sum(
                        float(e.get("amount", 0)) 
                        for e in entries 
                        if e.get("status", "").lower() in ["pending", "approved"]
                    )
                    break
        
        new_total = existing_month_total + submitted_month_total + amount
        
        print(f"\n💰 AMOUNT CALCULATION:")
        print(f"   Existing Temp Entries: ₹{existing_month_total}")
        print(f"   Submitted Entries: ₹{submitted_month_total}")
        print(f"   New Entry: +₹{amount}")
        print(f"   New Total: ₹{new_total}")
        print(f"   OPE Limit: ₹{actual_ope_limit}")
        
        # Check for duplicate entry
        if temp_doc:
            data_array = temp_doc.get("Data", [])
            for data_item in data_array:
                if formatted_month_range in data_item:
                    entries = data_item[formatted_month_range]
                    for entry in entries:
                        if (entry.get("date") == date and
                            entry.get("client") == client and
                            entry.get("project_id") == project_id and
                            entry.get("project_name") == project_name and
                            entry.get("project_type") == project_type and
                            entry.get("location_from") == location_from and
                            entry.get("location_to") == location_to and
                            entry.get("travel_mode") == travel_mode and
                            entry.get("amount") == amount):
                            
                            raise HTTPException(
                                status_code=400,
                                detail="⚠️ Duplicate Entry Detected!\n\nAn entry with the same details already exists for this date and month."
                            )
        
        # ============================================
        # Handle PDF - GridFS Upload
        # ============================================
        ticket_pdf_id = None
        if ticket_pdf:
            pdf_content = await ticket_pdf.read()
            pdf_filename = f"temp_ope_{employee_code}_{datetime.utcnow().timestamp()}.pdf"
            ticket_pdf_id = await upload_to_gridfs(pdf_content, pdf_filename)
            print(f"✅ PDF uploaded to GridFS: {ticket_pdf_id}")
        
        # Create entry
        entry_doc = {
            "_id": ObjectId(),
            "date": date,
            "client": client,
            "project_id": project_id,
            "project_name": project_name,
            "project_type": project_type,
            "location_from": location_from,
            "location_to": location_to,
            "travel_mode": travel_mode,
            "amount": amount,
            "remarks": remarks,
            "ticket_pdf": ticket_pdf_id,       # GridFS file_id or None
            "created_time": datetime.utcnow().isoformat(),
            "updated_time": datetime.utcnow().isoformat(),
            "status": "saved"
        }
        
        if not temp_doc:
            new_doc = {
                "employeeId": employee_code,
                "employeeName": emp.get("Emp Name", ""),
                "designation": emp.get("Designation Name", ""),
                "gender": emp.get("Gender", ""),
                "partner": emp.get("Partner", ""),
                "reportingManager": emp.get("ReportingEmpName", ""),
                "department": "",
                "Data": [
                    {
                        formatted_month_range: [entry_doc]
                    }
                ]
            }
            await db["Temp_OPE_data"].insert_one(new_doc)
            print(f"✅ NEW temp document created")
        else:
            month_exists = False
            data_array = temp_doc.get("Data", [])
            
            for i, data_item in enumerate(data_array):
                if formatted_month_range in data_item:
                    await db["Temp_OPE_data"].update_one(
                        {"employeeId": employee_code},
                        {"$push": {f"Data.{i}.{formatted_month_range}": entry_doc}}
                    )
                    month_exists = True
                    break
            
            if not month_exists:
                await db["Temp_OPE_data"].update_one(
                    {"employeeId": employee_code},
                    {"$push": {"Data": {formatted_month_range: [entry_doc]}}}
                )
        
        if new_total > actual_ope_limit:
            approval_levels = 3
            within_limit = False
            print(f"📊 New total (₹{new_total}) EXCEEDS limit → 3-level approval")
        else:
            approval_levels = 2
            within_limit = True
            print(f"📊 New total (₹{new_total}) WITHIN limit → 2-level approval")
        
        print(f"\n{'='*60}")
        print(f"✅ ENTRY SAVED SUCCESSFULLY")
        print(f"{'='*60}\n")
        
        return {
            "message": f"Entry saved successfully! {('✅ Within limit!' if within_limit else '⚠️ Exceeds limit!')}",
            "entry_id": str(entry_doc["_id"]),
            "status": "saved",
            "total_amount": new_total,
            "ope_limit": actual_ope_limit,
            "within_limit": within_limit,
            "approval_levels": approval_levels
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error saving temp entry: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ========================
# GET TEMPORARY HISTORY
# =========================
@app.get("/api/ope/temp-history/{employee_code}")
async def get_temp_history(employee_code: str, current_user=Depends(get_current_user)):
    try:
        print(f"📌 Fetching temp history for: {employee_code}")
        
        if current_user["employee_code"] != employee_code:
            raise HTTPException(status_code=403, detail="Access denied")
        
        temp_doc = await db["Temp_OPE_data"].find_one({"employeeId": employee_code})
        
        if not temp_doc:
            print(f"📭 No temp data found")
            return {"history": []}
        
        history = []
        data_array = temp_doc.get("Data", [])
        
        for data_item in data_array:
            for month_range, entries in data_item.items():
                for entry in entries:
                    history.append({
                        "_id": str(entry.get("_id", "")),
                        "month_range": month_range,
                        "date": entry.get("date"),
                        "client": entry.get("client"),
                        "project_id": entry.get("project_id"),
                        "project_name": entry.get("project_name"),
                        "project_type": entry.get("project_type", "N/A"),
                        "location_from": entry.get("location_from"),
                        "location_to": entry.get("location_to"),
                        "travel_mode": entry.get("travel_mode"),
                        "amount": entry.get("amount"),
                        "remarks": entry.get("remarks"),
                        "ticket_pdf": entry.get("ticket_pdf"),    # GridFS ID or None
                        "status": "saved"
                    })
        
        print(f"✅ Found {len(history)} temp entries")
        return {"history": history}
        
    except Exception as e:
        print(f"❌ Error fetching temp history: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ========================
# UPDATE TEMPORARY ENTRY
# ========================
@app.put("/api/ope/update-temp/{entry_id}")
async def update_temp_entry(
    entry_id: str,
    update_data: dict,
    current_user=Depends(get_current_user)
):
    try:
        employee_code = current_user["employee_code"]
        month_range = update_data.get("month_range")
        
        print(f"📝 Updating temp entry {entry_id} for: {employee_code}")
        
        if not month_range:
            raise HTTPException(status_code=400, detail="month_range required")
        
        temp_doc = await db["Temp_OPE_data"].find_one({"employeeId": employee_code})
        if not temp_doc:
            raise HTTPException(status_code=404, detail="No temporary data found")
        
        data_array = temp_doc.get("Data", [])
        updated = False
        
        for i, data_item in enumerate(data_array):
            if month_range in data_item:
                entries = data_item[month_range]
                for j, entry in enumerate(entries):
                    if str(entry.get("_id")) == entry_id:
                        update_fields = {
                            f"Data.{i}.{month_range}.{j}.date": update_data.get("date"),
                            f"Data.{i}.{month_range}.{j}.client": update_data.get("client"),
                            f"Data.{i}.{month_range}.{j}.project_id": update_data.get("project_id"),
                            f"Data.{i}.{month_range}.{j}.project_name": update_data.get("project_name"),
                            f"Data.{i}.{month_range}.{j}.project_type": update_data.get("project_type"),
                            f"Data.{i}.{month_range}.{j}.location_from": update_data.get("location_from"),
                            f"Data.{i}.{month_range}.{j}.location_to": update_data.get("location_to"),
                            f"Data.{i}.{month_range}.{j}.travel_mode": update_data.get("travel_mode"),
                            f"Data.{i}.{month_range}.{j}.amount": update_data.get("amount"),
                            f"Data.{i}.{month_range}.{j}.remarks": update_data.get("remarks"),
                            f"Data.{i}.{month_range}.{j}.updated_time": datetime.utcnow().isoformat()
                        }
                        
                        await db["Temp_OPE_data"].update_one(
                            {"employeeId": employee_code},
                            {"$set": update_fields}
                        )
                        updated = True
                        print(f"✅ Entry updated successfully")
                        break
            if updated:
                break
        
        if not updated:
            raise HTTPException(status_code=404, detail="Entry not found in temp data")
        
        return {"message": "Entry updated successfully"}
        
    except Exception as e:
        print(f"❌ Error updating temp entry: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# DELETE TEMPORARY ENTRY
# ============================================
@app.delete("/api/ope/delete-temp/{entry_id}")
async def delete_temp_entry(
    entry_id: str,
    delete_data: dict,
    current_user=Depends(get_current_user)
):
    try:
        employee_code = current_user["employee_code"]
        month_range = delete_data.get("month_range")
        
        print(f"🗑️ Deleting temp entry {entry_id} for: {employee_code}")
        
        if not entry_id or entry_id == "dummy":
            raise HTTPException(status_code=400, detail="Invalid entry ID")
        
        if not month_range:
            raise HTTPException(status_code=400, detail="month_range required")
        
        temp_doc = await db["Temp_OPE_data"].find_one({"employeeId": employee_code})
        if not temp_doc:
            raise HTTPException(status_code=404, detail="No temporary data found")
        
        data_array = temp_doc.get("Data", [])
        deleted = False
        
        for i, data_item in enumerate(data_array):
            if month_range in data_item:
                entries = data_item[month_range]
                
                for j, entry in enumerate(entries):
                    if str(entry.get("_id")) == entry_id:
                        print(f"✅ Found entry at Data.{i}.{month_range}.{j}")
                        
                        # Delete PDF from GridFS if exists
                        pdf_id = entry.get("ticket_pdf")
                        if pdf_id and is_gridfs_id(pdf_id):
                            await delete_from_gridfs(pdf_id)
                            print(f"✅ PDF deleted from GridFS: {pdf_id}")
                        
                        if len(entries) == 1:
                            print(f"🗑️ Removing entire month range")
                            await db["Temp_OPE_data"].update_one(
                                {"employeeId": employee_code},
                                {"$pull": {"Data": {month_range: {"$exists": True}}}}
                            )
                        else:
                            print(f"🗑️ Removing single entry")
                            await db["Temp_OPE_data"].update_one(
                                {"employeeId": employee_code},
                                {"$pull": {f"Data.{i}.{month_range}": {"_id": ObjectId(entry_id)}}}
                            )
                        
                        deleted = True
                        break
            
            if deleted:
                break
        
        if not deleted:
            raise HTTPException(status_code=404, detail="Entry not found in temp data")
        
        print(f"✅ Entry deleted successfully")
        return {
            "message": "Entry deleted successfully",
            "entry_id": entry_id
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error deleting temp entry: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

            
# ============================================
# SUBMIT ALL ENTRIES - COMPLETE UPDATED VERSION
# ============================================
@app.post("/api/ope/submit-final")
async def submit_final_entries(
    request: Request,
    current_user=Depends(get_current_user)
):
    try:
        employee_code = current_user["employee_code"].strip().upper()
        
        body = await request.json()
        month_range = body.get("month_range")
        
        print(f"🚀 SUBMIT FINAL: Employee {employee_code}, Month {month_range}")
        
        if not month_range:
            raise HTTPException(status_code=400, detail="month_range required")
        
        def format_month_range(month_str):
            try:
                parts = month_str.lower().split('-')
                month_map = {
                    'jan': 'Jan', 'feb': 'Feb', 'mar': 'Mar', 'apr': 'Apr',
                    'may': 'May', 'jun': 'Jun', 'jul': 'Jul', 'aug': 'Aug',
                    'sep': 'Sep', 'oct': 'Oct', 'nov': 'Nov', 'dec': 'Dec'
                }
                
                if len(parts) == 3:
                    month1 = month_map.get(parts[0], parts[0].capitalize())
                    month2 = month_map.get(parts[1], parts[1].capitalize())
                    year = parts[2]
                    return f"{month1} {year} - {month2} {year}"
                elif len(parts) == 2:
                    month = month_map.get(parts[0], parts[0].capitalize())
                    year = parts[1]
                    return f"{month} {year}"
                else:
                    return month_str
            except Exception as e:
                return month_str
        
        formatted_month_range = format_month_range(month_range)
        print(f"📅 Formatted month range: {formatted_month_range}")
        
        temp_doc = await db["Temp_OPE_data"].find_one({"employeeId": employee_code})
        
        if not temp_doc:
            raise HTTPException(status_code=404, detail="No temporary data found to submit")
        
        entries_to_submit = []
        data_array = temp_doc.get("Data", [])
        
        for data_item in data_array:
            if formatted_month_range in data_item:
                entries_to_submit = data_item[formatted_month_range]
                break
        
        if not entries_to_submit:
            raise HTTPException(status_code=404, detail=f"No entries found for {formatted_month_range}")
        
        print(f"📦 Found {len(entries_to_submit)} entries to submit")
        
        emp = await db["Employee_details"].find_one({"EmpID": employee_code})
        if not emp:
            raise HTTPException(status_code=404, detail="Employee details not found")
        
        # Check if submitter is a Reporting Manager
        is_reporting_manager = await db["Reporting_managers"].find_one({"ReportingEmpCode": employee_code})
        
        reporting_manager_code = emp.get("ReportingEmpCode", "").strip().upper()
        reporting_manager_name = emp.get("ReportingEmpName", "")
        partner_code = emp.get("PartnerEmpCode", "").strip().upper()
        partner_name = emp.get("Partner", "")
        
        new_entries_amount = sum(float(entry.get("amount", 0)) for entry in entries_to_submit)
        
        status_doc = await db["Status"].find_one({"employeeId": employee_code})
        
        existing_total = 0
        month_exists = False
        existing_month_index = -1
        
        if status_doc:
            approval_status = status_doc.get("approval_status", [])
            for i, ps in enumerate(approval_status):
                if ps.get("payroll_month") == formatted_month_range:
                    existing_total = ps.get("total_amount", 0)
                    month_exists = True
                    existing_month_index = i
                    print(f"📊 Found existing month entry with total: ₹{existing_total}")
                    break
        
        cumulative_total = existing_total + new_entries_amount
        
        current_time = datetime.utcnow().isoformat()
        
        # ============================================
        # APPROVAL FLOW LOGIC BASED ON SUBMITTER TYPE
        # ============================================
        if is_reporting_manager:
            print(f"\n{'='*60}")
            print(f"👔 SUBMITTER IS A REPORTING MANAGER")
            print(f"   Employee Code: {employee_code}")
            print(f"   Partner: {partner_code} ({partner_name})")
            print(f"{'='*60}\n")
            
            if not partner_code:
                raise HTTPException(status_code=400, detail="No Partner assigned to this Reporting Manager")
            
            total_levels = 2
            ope_label = "Reporting_Manager"
            
            payroll_entry = {
                "payroll_month": formatted_month_range,
                "ope_label": ope_label,
                "submitter_type": "Reporting_Manager",
                "total_levels": total_levels,
                "limit": 0,
                "total_amount": cumulative_total,
                "L1": {
                    "status": False,
                    "approver_name": partner_name,
                    "approver_code": partner_code,
                    "approved_date": None,
                    "level_name": "Partner"
                },
                "L2": {
                    "status": False,
                    "approver_name": "HR",
                    "approver_code": "JHS729",
                    "approved_date": None,
                    "level_name": "HR"
                },
                "current_level": "L1",
                "overall_status": "pending",
                "submission_date": current_time
            }
            
            pending_approver_code = partner_code
            
        else:
            print(f"\n{'='*60}")
            print(f"👤 SUBMITTER IS A REGULAR EMPLOYEE")
            print(f"   Employee Code: {employee_code}")
            print(f"   Reporting Manager: {reporting_manager_code} ({reporting_manager_name})")
            print(f"{'='*60}\n")
            
            if not reporting_manager_code:
                raise HTTPException(status_code=400, detail="No reporting manager assigned")
            
            ope_limit = emp.get("OPE LIMIT")
            
            if ope_limit is None:
                ope_limit = 1500
                print(f"⚠️ OPE Limit not found in Employee_details, using default: ₹{ope_limit}")
            else:
                ope_limit = float(ope_limit)
                print(f"✅ OPE Limit from Employee_details: ₹{ope_limit}")
            
            print(f"\n{'='*60}")
            print(f"💰 AMOUNT CALCULATION:")
            print(f"   Previous Total: ₹{existing_total}")
            print(f"   New Entries: +₹{new_entries_amount}")
            print(f"   Cumulative Total: ₹{cumulative_total}")
            print(f"   OPE Limit: ₹{ope_limit}")
            print(f"{'='*60}\n")
            
            if cumulative_total > ope_limit:
                ope_label = "Greater"
                total_levels = 3
                print(f"📊 Cumulative amount (₹{cumulative_total}) EXCEEDS limit (₹{ope_limit}) → 3-level approval required")
            else:
                ope_label = "Less"
                total_levels = 2
                print(f"📊 Cumulative amount (₹{cumulative_total}) WITHIN limit (₹{ope_limit}) → 2-level approval required")
            
            payroll_entry = {
                "payroll_month": formatted_month_range,
                "ope_label": ope_label,
                "submitter_type": "Employee",
                "total_levels": total_levels,
                "limit": ope_limit,
                "total_amount": cumulative_total,
                "L1": {
                    "status": False,
                    "approver_name": reporting_manager_name,
                    "approver_code": reporting_manager_code,
                    "approved_date": None,
                    "level_name": "Reporting Manager"
                },
                "L2": {
                    "status": False,
                    "approver_name": "HR" if total_levels == 2 else partner_name,
                    "approver_code": "JHS729" if total_levels == 2 else partner_code,
                    "approved_date": None,
                    "level_name": "HR" if total_levels == 2 else "Partner"
                },
                "current_level": "L1",
                "overall_status": "pending",
                "submission_date": current_time
            }
            
            if total_levels == 3:
                payroll_entry["L3"] = {
                    "status": False,
                    "approver_name": "HR",
                    "approver_code": "JHS729",
                    "approved_date": None,
                    "level_name": "HR"
                }
                print(f"✅ Added L3 (HR) level for approval")
            
            pending_approver_code = reporting_manager_code
        
        # ============================================
        # CREATE OR UPDATE STATUS DOCUMENT
        # ============================================
        if not status_doc:
            new_status_doc = {
                "employeeId": employee_code,
                "employeeName": emp.get("Emp Name", ""),
                "ReportingEmpCode": reporting_manager_code, 
                "PartnerEmpCode": partner_code,  
                "HREmpCode": "JHS729", 
                "approval_status": [payroll_entry]
            }
            result = await db["Status"].insert_one(new_status_doc)
            status_doc_id = str(result.inserted_id)
            
        else:
            status_doc_id = str(status_doc["_id"])
            
            if month_exists:
                print(f"🔄 Updating existing month entry at index {existing_month_index}")
                
                update_fields = {
                    f"approval_status.{existing_month_index}.total_amount": cumulative_total,
                    f"approval_status.{existing_month_index}.ope_label": ope_label,
                    f"approval_status.{existing_month_index}.total_levels": total_levels,
                    f"approval_status.{existing_month_index}.submitter_type": payroll_entry["submitter_type"],
                    f"approval_status.{existing_month_index}.submission_date": current_time
                }
                
                if is_reporting_manager:
                    update_fields[f"approval_status.{existing_month_index}.limit"] = 0
                else:
                    update_fields[f"approval_status.{existing_month_index}.limit"] = ope_limit
                
                if total_levels == 3:
                    update_fields[f"approval_status.{existing_month_index}.L3"] = {
                        "status": False,
                        "approver_name": "HR",
                        "approver_code": "JHS729",
                        "approved_date": None,
                        "level_name": "HR"
                    }
                    update_fields[f"approval_status.{existing_month_index}.L2.approver_name"] = partner_name
                    update_fields[f"approval_status.{existing_month_index}.L2.approver_code"] = partner_code
                    update_fields[f"approval_status.{existing_month_index}.L2.level_name"] = "Partner"
                else:
                    if is_reporting_manager:
                        update_fields[f"approval_status.{existing_month_index}.L2.approver_name"] = "HR"
                        update_fields[f"approval_status.{existing_month_index}.L2.approver_code"] = "JHS729"
                        update_fields[f"approval_status.{existing_month_index}.L2.level_name"] = "HR"
                    else:
                        update_fields[f"approval_status.{existing_month_index}.L2.approver_name"] = "HR"
                        update_fields[f"approval_status.{existing_month_index}.L2.approver_code"] = "JHS729"
                        update_fields[f"approval_status.{existing_month_index}.L2.level_name"] = "HR"
                
                await db["Status"].update_one(
                    {"employeeId": employee_code},
                    {"$set": update_fields}
                )
                
                print(f"✅ Updated existing payroll month with cumulative total: ₹{cumulative_total}")
                
            else:
                await db["Status"].update_one(
                    {"employeeId": employee_code},
                    {"$push": {"approval_status": payroll_entry}}
                )
                print(f"✅ Added new payroll month: {formatted_month_range}")
        
        # Update each entry with status reference
        for entry in entries_to_submit:
            entry["status"] = "pending"
            entry["submitted_time"] = current_time
            entry["status_doc_id"] = status_doc_id
            entry["payroll_month"] = formatted_month_range
            entry["approved_by"] = None
            entry["approved_date"] = None
            entry["rejected_by"] = None
            entry["rejected_date"] = None
            entry["rejection_reason"] = None
        
        # Move to OPE_data collection
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_code})
        
        if not ope_doc:
            new_doc = {
                "employeeId": employee_code,
                "employeeName": emp.get("Emp Name", ""),
                "designation": emp.get("Designation Name", ""),
                "gender": emp.get("Gender", ""),
                "partner": emp.get("Partner", ""),
                "reportingManager": emp.get("ReportingEmpName", ""),
                "department": "",
                "Data": [
                    {
                        formatted_month_range: entries_to_submit
                    }
                ]
            }
            await db["OPE_data"].insert_one(new_doc)
            print(f"✅ Created new OPE_data document")
        else:
            month_exists_in_ope = False
            ope_data_array = ope_doc.get("Data", [])
            
            for i, data_item in enumerate(ope_data_array):
                if formatted_month_range in data_item:
                    for entry in entries_to_submit:
                        await db["OPE_data"].update_one(
                            {"employeeId": employee_code},
                            {"$push": {f"Data.{i}.{formatted_month_range}": entry}}
                        )
                    month_exists_in_ope = True
                    print(f"✅ Appended to existing month in OPE_data")
                    break
            
            if not month_exists_in_ope:
                await db["OPE_data"].update_one(
                    {"employeeId": employee_code},
                    {"$push": {"Data": {formatted_month_range: entries_to_submit}}}
                )
                print(f"✅ Added new month range to OPE_data")
        
        # Add to PENDING collection
        pending_doc = await db["Pending"].find_one({"ReportingEmpCode": pending_approver_code})
        
        if not pending_doc:
            await db["Pending"].insert_one({
                "ReportingEmpCode": pending_approver_code,
                "EmployeesCodes": [employee_code]
            })
            print(f"✅ Created NEW Pending document for approver {pending_approver_code}")
        else:
            if employee_code not in pending_doc.get("EmployeesCodes", []):
                await db["Pending"].update_one(
                    {"ReportingEmpCode": pending_approver_code},
                    {"$addToSet": {"EmployeesCodes": employee_code}}
                )
                print(f"✅ Added employee to Pending list under {pending_approver_code}")
        
        # Delete from Temp_OPE_data
        temp_data_array = temp_doc.get("Data", [])
        
        for i, data_item in enumerate(temp_data_array):
            if formatted_month_range in data_item:
                await db["Temp_OPE_data"].update_one(
                    {"employeeId": employee_code},
                    {"$pull": {"Data": {formatted_month_range: {"$exists": True}}}}
                )
                print(f"✅ Removed from Temp_OPE_data")
                break
        
        updated_temp = await db["Temp_OPE_data"].find_one({"employeeId": employee_code})
        if updated_temp and len(updated_temp.get("Data", [])) == 0:
            await db["Temp_OPE_data"].delete_one({"employeeId": employee_code})
            print(f"✅ Deleted empty Temp_OPE_data document")
        
        print(f"\n{'='*60}")
        print(f"✅✅ SUBMISSION COMPLETE ✅✅")
        print(f"   Submitter Type: {'REPORTING MANAGER' if is_reporting_manager else 'EMPLOYEE'}")
        print(f"   Employee: {employee_code}")
        print(f"   Previous Total: ₹{existing_total}")
        print(f"   New Entries: +₹{new_entries_amount}")
        print(f"   Cumulative Total: ₹{cumulative_total}")
        print(f"   Approval Levels: {total_levels}")
        print(f"   First Approver: {pending_approver_code}")
        print(f"{'='*60}\n")
        
        return {
            "message": "Entries submitted successfully for approval",
            "submitted_count": len(entries_to_submit),
            "month_range": formatted_month_range,
            "submitter_type": "Reporting_Manager" if is_reporting_manager else "Employee",
            "first_approver": pending_approver_code,
            "previous_total": existing_total,
            "new_entries_amount": new_entries_amount,
            "total_amount": cumulative_total,
            "ope_limit": 0 if is_reporting_manager else ope_limit,
            "ope_label": ope_label,
            "total_levels": total_levels,
            "status": "pending_approval"
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error submitting final: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))    


@app.post("/api/ope/manager/approve/{employee_code}")
async def approve_employee_entries(
    employee_code: str,
    request: Request,
    current_user=Depends(get_current_user)
):
    try:
        reporting_emp_code = current_user["employee_code"].strip().upper()
        employee_code = employee_code.strip().upper()
        
        body = await request.json()
        approval_remark = body.get("remark", "Approved without remark")
        
        print(f"\n{'='*60}")
        print(f"✅ MANAGER APPROVAL REQUEST")
        print(f"Manager: {reporting_emp_code}")
        print(f"Employee: {employee_code}")
        print(f"Remark: {approval_remark}")
        print(f"{'='*60}\n")
        
        manager = await db["Reporting_managers"].find_one({"ReportingEmpCode": reporting_emp_code})
        if not manager:
            raise HTTPException(status_code=403, detail="You are not a reporting manager")
        
        manager_name = manager.get("ReportingEmpName", reporting_emp_code)
        
        emp = await db["Employee_details"].find_one({"EmpID": employee_code})
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        
        emp_reporting_manager_code = emp.get("ReportingEmpCode", "").strip().upper()
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_code})
        if not ope_doc:
            raise HTTPException(status_code=404, detail="No OPE data found for employee")
        
        approved_count = 0
        current_time = datetime.utcnow().isoformat()
        
        data_array = ope_doc.get("Data", [])
        payroll_months_approved = set()
        
        for i, data_item in enumerate(data_array):
            for month_range, entries in data_item.items():
                for j, entry in enumerate(entries):
                    entry_status = entry.get("status", "").lower()
                    
                    if entry_status != "pending":
                        continue
                    
                    await db["OPE_data"].update_one(
                        {"employeeId": employee_code},
                        {"$set": {
                            f"Data.{i}.{month_range}.{j}.status": "approved",
                            f"Data.{i}.{month_range}.{j}.approved_date": current_time,
                            f"Data.{i}.{month_range}.{j}.approved_by": reporting_emp_code,
                            f"Data.{i}.{month_range}.{j}.approver_name": manager_name,
                            f"Data.{i}.{month_range}.{j}.approval_remark": approval_remark,
                            f"Data.{i}.{month_range}.{j}.L1_approved": True,
                            f"Data.{i}.{month_range}.{j}.L1_approver_code": reporting_emp_code,
                            f"Data.{i}.{month_range}.{j}.L1_approver_name": manager_name
                        }}
                    )
                    
                    payroll_months_approved.add(month_range)
                    approved_count += 1
        
        if approved_count == 0:
            raise HTTPException(status_code=404, detail="No pending entries found")
        
        status_doc = await db["Status"].find_one({"employeeId": employee_code})
        
        partner_code = None
        
        if status_doc:
            approval_status = status_doc.get("approval_status", [])
            
            for i, ps in enumerate(approval_status):
                if ps.get("payroll_month") in payroll_months_approved:
                    total_levels = ps.get("total_levels", 2)
                    submitter_type = ps.get("submitter_type", "Employee")
                    
                    print(f"\n📊 Processing payroll: {ps.get('payroll_month')}")
                    print(f"   Total Levels: {total_levels}")
                    print(f"   Submitter Type: {submitter_type}")
                    
                    if total_levels == 2:
                        await db["Status"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"approval_status.{i}.L1.status": True,
                                f"approval_status.{i}.L1.approver_code": reporting_emp_code,
                                f"approval_status.{i}.L1.approver_name": manager_name,
                                f"approval_status.{i}.L1.approved_date": current_time,
                                f"approval_status.{i}.L1.approval_remark": approval_remark,
                                f"approval_status.{i}.overall_status": "pending",
                                f"approval_status.{i}.current_level": "L2"
                            }}
                        )
                        print(f"   ✅ 2-level: L1 approved → L2 (HR) pending")
                        
                    elif total_levels == 3:
                        partner_code = ps.get("L2", {}).get("approver_code")
                        
                        if not partner_code:
                            partner_code = emp.get("PartnerEmpCode", "").strip().upper()
                        
                        print(f"   🔥 3-level: L1 approved → L2 (Partner: {partner_code}) pending")
                        
                        await db["Status"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"approval_status.{i}.L1.status": True,
                                f"approval_status.{i}.L1.approver_code": reporting_emp_code,
                                f"approval_status.{i}.L1.approver_name": manager_name,
                                f"approval_status.{i}.L1.approved_date": current_time,
                                f"approval_status.{i}.L1.approval_remark": approval_remark,
                                f"approval_status.{i}.overall_status": "pending",
                                f"approval_status.{i}.current_level": "L2"
                            }}
                        )
        
        await db["Pending"].update_one(
            {"ReportingEmpCode": emp_reporting_manager_code},
            {"$pull": {"EmployeesCodes": employee_code}}
        )
        print(f"✅ Removed from Manager's Pending")
        
        approved_doc = await db["Approved"].find_one({"ReportingEmpCode": reporting_emp_code})
        if not approved_doc:
            await db["Approved"].insert_one({
                "ReportingEmpCode": reporting_emp_code,
                "EmployeesCodes": [employee_code]
            })
            print(f"✅ Created NEW Approved document")
        else:
            if employee_code not in approved_doc.get("EmployeesCodes", []):
                await db["Approved"].update_one(
                    {"ReportingEmpCode": reporting_emp_code},
                    {"$addToSet": {"EmployeesCodes": employee_code}}
                )
                print(f"✅ Added to Manager's Approved collection")
        
        if partner_code:
            print(f"\n🔥 ROUTING TO PARTNER: {partner_code}")
            
            partner_pending_doc = await db["Pending"].find_one({"ReportingEmpCode": partner_code})
            
            if not partner_pending_doc:
                await db["Pending"].insert_one({
                    "ReportingEmpCode": partner_code,
                    "EmployeesCodes": [employee_code]
                })
                print(f"   ✅ Created NEW Pending document for Partner {partner_code}")
            else:
                if employee_code not in partner_pending_doc.get("EmployeesCodes", []):
                    await db["Pending"].update_one(
                        {"ReportingEmpCode": partner_code},
                        {"$addToSet": {"EmployeesCodes": employee_code}}
                    )
                    print(f"   ✅ Added to Partner's Pending collection")
        
        print(f"\n✅✅ APPROVAL COMPLETE")
        print(f"   Total approved: {approved_count}")
        print(f"   Approval Remark: {approval_remark}")
        print(f"   Next level: {'L2 (Partner)' if partner_code else 'L2 (HR)'}")
        print(f"{'='*60}\n")
        
        return {
            "message": f"Approved {approved_count} entries",
            "approved_count": approved_count,
            "approval_remark": approval_remark,
            "employee_code": employee_code,
            "next_level": "L2",
            "next_approver": partner_code if partner_code else "HR"
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"\n❌❌ ERROR:")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/ope/manager/approved-list")
async def get_approved_employees_list(current_user=Depends(get_current_user)):
    try:
        reporting_emp_code = current_user["employee_code"].strip().upper()
        
        print(f"📋 Fetching approved list for manager: {reporting_emp_code}")
        
        manager = await db["Reporting_managers"].find_one({"ReportingEmpCode": reporting_emp_code})
        if not manager:
            raise HTTPException(status_code=403, detail="You are not a reporting manager")
        
        approved_doc = await db["Approved"].find_one({"ReportingEmpCode": reporting_emp_code})
        
        employee_codes = []
        if approved_doc:
            employee_codes = approved_doc.get("EmployeesCodes", [])
        
        print(f"✅ Found {len(employee_codes)} approved employees")
        
        return {
            "reporting_manager": reporting_emp_code,
            "employee_codes": employee_codes,
            "count": len(employee_codes)
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/ope/manager/rejected-list")
async def get_rejected_employees_list(current_user=Depends(get_current_user)):
    try:
        reporting_emp_code = current_user["employee_code"].strip().upper()
        
        print(f"📋 Fetching rejected list for manager: {reporting_emp_code}")
        
        manager = await db["Reporting_managers"].find_one({"ReportingEmpCode": reporting_emp_code})
        if not manager:
            raise HTTPException(status_code=403, detail="You are not a reporting manager")
        
        rejected_doc = await db["Rejected"].find_one({"ReportingEmpCode": reporting_emp_code})
        
        employee_codes = []
        if rejected_doc:
            employee_codes = rejected_doc.get("EmployeesCodes", [])
        
        print(f"✅ Found {len(employee_codes)} rejected employees")
        
        return {
            "reporting_manager": reporting_emp_code,
            "employee_codes": employee_codes,
            "count": len(employee_codes)
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ope/manager/reject-single")
async def reject_single_entry(
    request: Request,
    current_user=Depends(get_current_user)
):
    try:
        reporting_emp_code = current_user["employee_code"].strip().upper()
        
        body = await request.json()
        entry_id = body.get("entry_id")
        employee_id = body.get("employee_id")
        reason = body.get("reason", "No reason provided")
        
        print(f"❌ Rejecting entry {entry_id} for employee {employee_id}")
        
        manager = await db["Reporting_managers"].find_one({"ReportingEmpCode": reporting_emp_code})
        if not manager:
            raise HTTPException(status_code=403, detail="You are not a reporting manager")
        
        manager_name = manager.get("ReportingEmpName", reporting_emp_code)
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_id})
        if not ope_doc:
            raise HTTPException(status_code=404, detail="Employee data not found")
        
        data_array = ope_doc.get("Data", [])
        updated = False
        current_time = datetime.utcnow().isoformat()
        
        for i, data_item in enumerate(data_array):
            for month_range, entries in data_item.items():
                for j, entry in enumerate(entries):
                    if str(entry.get("_id")) == entry_id and entry.get("status") == "approved":
                        await db["OPE_data"].update_one(
                            {"employeeId": employee_id},
                            {"$set": {
                                f"Data.{i}.{month_range}.{j}.status": "rejected",
                                f"Data.{i}.{month_range}.{j}.rejected_by": reporting_emp_code,
                                f"Data.{i}.{month_range}.{j}.rejector_name": manager_name,
                                f"Data.{i}.{month_range}.{j}.rejected_date": current_time,
                                f"Data.{i}.{month_range}.{j}.rejection_reason": reason
                            }}
                        )
                        
                        status_id = entry.get("status_id")
                        if status_id:
                            await db["Status"].update_one(
                                {"_id": ObjectId(status_id)},
                                {"$set": {
                                    "overall_status": "rejected",
                                    "L1.status": False,
                                    "L1.rejected_by": reporting_emp_code,
                                    "L1.rejected_date": current_time
                                }}
                            )
                        
                        updated = True
                        break
            if updated:
                break
        
        if not updated:
            raise HTTPException(status_code=404, detail="Entry not found or not approved")
        
        all_rejected = True
        for i, data_item in enumerate(data_array):
            for month_range, entries in data_item.items():
                for entry in entries:
                    if str(entry.get("_id")) != entry_id and entry.get("status") == "approved":
                        all_rejected = False
                        break
        
        if all_rejected:
            print(f"🔄 Moving employee from Approved → Rejected")
            
            await db["Approved"].update_one(
                {"ReportingEmpCode": reporting_emp_code},
                {"$pull": {"EmployeesCodes": employee_id}}
            )
            print(f"✅ Removed from Approved")
            
            rejected_doc = await db["Rejected"].find_one({"ReportingEmpCode": reporting_emp_code})
            if not rejected_doc:
                await db["Rejected"].insert_one({
                    "ReportingEmpCode": reporting_emp_code,
                    "EmployeesCodes": [employee_id]
                })
                print(f"✅ Created NEW Rejected document")
            else:
                if employee_id not in rejected_doc.get("EmployeesCodes", []):
                    await db["Rejected"].update_one(
                        {"ReportingEmpCode": reporting_emp_code},
                        {"$addToSet": {"EmployeesCodes": employee_id}}
                    )
                    print(f"✅ Added to Rejected collection")
        
        return {"message": "Entry rejected successfully"}
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    

@app.post("/api/ope/manager/approve-single")
async def approve_single_entry(
    request: Request,
    current_user=Depends(get_current_user)
):
    try:
        reporting_emp_code = current_user["employee_code"].strip().upper()
        
        body = await request.json()
        entry_id = body.get("entry_id")
        employee_id = body.get("employee_id")
        
        print(f"\n{'='*60}")
        print(f"✅ APPROVING REJECTED ENTRY")
        print(f"Manager: {reporting_emp_code}")
        print(f"Employee: {employee_id}")
        print(f"Entry ID: {entry_id}")
        print(f"{'='*60}\n")
        
        manager = await db["Reporting_managers"].find_one({"ReportingEmpCode": reporting_emp_code})
        if not manager:
            raise HTTPException(status_code=403, detail="You are not a reporting manager")
        
        manager_name = manager.get("ReportingEmpName", reporting_emp_code)
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_id})
        if not ope_doc:
            raise HTTPException(status_code=404, detail="Employee data not found")
        
        data_array = ope_doc.get("Data", [])
        updated = False
        current_time = datetime.utcnow().isoformat()
        
        for i, data_item in enumerate(data_array):
            for month_range, entries in data_item.items():
                for j, entry in enumerate(entries):
                    if str(entry.get("_id")) == entry_id and entry.get("status") == "rejected":
                        await db["OPE_data"].update_one(
                            {"employeeId": employee_id},
                            {"$set": {
                                f"Data.{i}.{month_range}.{j}.status": "approved",
                                f"Data.{i}.{month_range}.{j}.approved_by": reporting_emp_code,
                                f"Data.{i}.{month_range}.{j}.approver_name": manager_name,
                                f"Data.{i}.{month_range}.{j}.approved_date": current_time,
                                f"Data.{i}.{month_range}.{j}.rejected_by": None,
                                f"Data.{i}.{month_range}.{j}.rejector_name": None,
                                f"Data.{i}.{month_range}.{j}.rejected_date": None,
                                f"Data.{i}.{month_range}.{j}.rejection_reason": None
                            }}
                        )
                        
                        status_id = entry.get("status_id")
                        if status_id:
                            await db["Status"].update_one(
                                {"_id": ObjectId(status_id)},
                                {"$set": {
                                    "overall_status": "approved",
                                    "L1.status": True,
                                    "L1.approver_code": reporting_emp_code,
                                    "L1.approver_name": manager_name,
                                    "L1.approved_date": current_time,
                                    "L1.rejected_by": None,
                                    "L1.rejected_date": None
                                }}
                            )
                        
                        updated = True
                        print(f"✅ Entry updated to approved")
                        break
            if updated:
                break
        
        if not updated:
            raise HTTPException(status_code=404, detail="Entry not found or not rejected")
        
        all_approved = True
        any_rejected = False
        
        for i, data_item in enumerate(data_array):
            for month_range, entries in data_item.items():
                for entry in entries:
                    entry_status = entry.get("status", "").lower()
                    if str(entry.get("_id")) != entry_id:
                        if entry_status == "rejected":
                            any_rejected = True
                            all_approved = False
                            break
                        elif entry_status != "approved":
                            all_approved = False
        
        print(f"📊 Status check:")
        print(f"   All approved: {all_approved}")
        print(f"   Any rejected: {any_rejected}")
        
        if not any_rejected:
            print(f"🔄 Moving employee from Rejected → Approved")
            
            await db["Rejected"].update_one(
                {"ReportingEmpCode": reporting_emp_code},
                {"$pull": {"EmployeesCodes": employee_id}}
            )
            print(f"✅ Removed from Rejected collection")
            
            approved_doc = await db["Approved"].find_one({"ReportingEmpCode": reporting_emp_code})
            if not approved_doc:
                await db["Approved"].insert_one({
                    "ReportingEmpCode": reporting_emp_code,
                    "EmployeesCodes": [employee_id]
                })
                print(f"✅ Created NEW Approved document")
            else:
                if employee_id not in approved_doc.get("EmployeesCodes", []):
                    await db["Approved"].update_one(
                        {"ReportingEmpCode": reporting_emp_code},
                        {"$addToSet": {"EmployeesCodes": employee_id}}
                    )
                    print(f"✅ Added to Approved collection")
                else:
                    print(f"⚠️ Employee already in Approved collection")
        else:
            print(f"⚠️ Employee still has rejected entries, not moving collections")
        
        print(f"{'='*60}\n")
        
        return {"message": "Entry approved successfully"}
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
        
@app.get("/api/ope/status/{employee_code}")
async def get_employee_status(employee_code: str, current_user=Depends(get_current_user)):
    try:
        employee_code = employee_code.strip().upper()
        current_emp_code = current_user["employee_code"].strip().upper()
        
        print(f"\n{'='*60}")
        print(f"📊 FETCHING STATUS FOR: {employee_code}")
        print(f"{'='*60}\n")

        is_hr = (current_emp_code == "JHS729")
        is_manager = await db["Reporting_managers"].find_one({"ReportingEmpCode": current_emp_code})
        is_partner = await db["Partner"].find_one({"PartnerEmpCode": current_emp_code})

        if current_emp_code != employee_code and not is_hr and not is_manager and not is_partner:
            raise HTTPException(status_code=403, detail="Access denied")
        
        status_doc = await db["Status"].find_one({"employeeId": employee_code})
        
        if not status_doc:
            print(f"📭 No status document found for {employee_code}")
            return {"status_entries": []}
        
        approval_status = status_doc.get("approval_status", [])
        
        print(f"✅ Found {len(approval_status)} payroll months")
        
        status_entries = []
        
        for ps_index, ps in enumerate(approval_status):
            print(f"\n{'─'*40}")
            print(f"📅 Processing payroll: {ps.get('payroll_month')}")
            print(f"{'─'*40}")
            
            L1 = ps.get("L1", {})
            L2 = ps.get("L2", {})
            L3 = ps.get("L3", {}) if "L3" in ps else {}
            
            total_levels = ps.get("total_levels", 2)
            current_level = ps.get("current_level", "L1")
            overall_status = ps.get("overall_status", "pending")
            
            is_rejected = False
            rejected_level = None
            rejected_by = None
            rejected_by_name = None
            rejected_date = None
            rejection_reason = None
            
            l1_rejected = L1.get("rejected", False)
            if l1_rejected or L1.get("rejected_by"):
                is_rejected = True
                rejected_level = "L1"
                rejected_by = L1.get("rejected_by")
                rejected_by_name = L1.get("rejector_name") or L1.get("rejected_by")
                rejected_date = L1.get("rejected_date")
                rejection_reason = L1.get("rejection_reason")
                overall_status = "rejected"
                print(f"   ❌ L1 REJECTED")
            
            elif L2.get("rejected", False) or L2.get("rejected_by"):
                is_rejected = True
                rejected_level = "L2"
                rejected_by = L2.get("rejected_by")
                rejected_by_name = L2.get("rejector_name") or L2.get("rejected_by")
                rejected_date = L2.get("rejected_date")
                rejection_reason = L2.get("rejection_reason")
                overall_status = "rejected"
                print(f"   ❌ L2 REJECTED")
            
            elif L3 and (L3.get("rejected", False) or L3.get("rejected_by")):
                is_rejected = True
                rejected_level = "L3"
                rejected_by = L3.get("rejected_by")
                rejected_by_name = L3.get("rejector_name") or L3.get("rejected_by")
                rejected_date = L3.get("rejected_date")
                rejection_reason = L3.get("rejection_reason")
                overall_status = "rejected"
                print(f"   ❌ L3 REJECTED")
            
            entry = {
                "employeeId": employee_code,
                "employeeName": status_doc.get("employeeName", ""),
                "payroll_month": ps.get("payroll_month"),
                "ope_label": ps.get("ope_label"),
                "total_levels": total_levels,
                "limit": ps.get("limit"),
                "total_amount": ps.get("total_amount"),
                "original_total": ps.get("original_total", ps.get("total_amount")),
                "last_edited_by": ps.get("last_edited_by"),
                "last_edited_by_name": ps.get("last_edited_by_name"),
                "last_edited_by_role": ps.get("last_edited_by_role"),
                "last_edited_date": ps.get("last_edited_date"),
                "amount_edit_history": ps.get("amount_edit_history", []),
                "L1": {
                    "status": L1.get("status", False),
                    "level_name": L1.get("level_name", "Reporting Manager"),
                    "approver_code": L1.get("approver_code"),
                    "approver_name": L1.get("approver_name"),
                    "approved_date": L1.get("approved_date"),
                    "approval_remark": L1.get("approval_remark"),
                    "rejected": L1.get("rejected", False),
                    "rejected_by": L1.get("rejected_by"),
                    "rejector_name": L1.get("rejector_name"),
                    "rejected_date": L1.get("rejected_date"),
                    "rejection_reason": L1.get("rejection_reason")
                },
                "L2": {
                    "status": L2.get("status", False),
                    "level_name": L2.get("level_name", "Partner" if total_levels == 3 else "HR"),
                    "approver_code": L2.get("approver_code"),
                    "approver_name": L2.get("approver_name"),
                    "approved_date": L2.get("approved_date"),
                    "approval_remark": L2.get("approval_remark"),
                    "rejected": L2.get("rejected", False),
                    "rejected_by": L2.get("rejected_by"),
                    "rejector_name": L2.get("rejector_name"),
                    "rejected_date": L2.get("rejected_date"),
                    "rejection_reason": L2.get("rejection_reason")
                },
                "current_level": current_level,
                "overall_status": overall_status,
                "submission_date": ps.get("submission_date"),
                "is_rejected": is_rejected,
                "rejected_level": rejected_level,
                "rejected_by": rejected_by,
                "rejected_by_name": rejected_by_name,
                "rejected_date": rejected_date,
                "rejection_reason": rejection_reason
            }
            
            if L3:
                entry["L3"] = {
                    "status": L3.get("status", False),
                    "level_name": L3.get("level_name", "HR"),
                    "approver_code": L3.get("approver_code"),
                    "approver_name": L3.get("approver_name"),
                    "approved_date": L3.get("approved_date"),
                    "approval_remark": L3.get("approval_remark"),
                    "rejected": L3.get("rejected", False),
                    "rejected_by": L3.get("rejected_by"),
                    "rejector_name": L3.get("rejector_name"),
                    "rejected_date": L3.get("rejected_date"),
                    "rejection_reason": L3.get("rejection_reason")
                }
            
            status_entries.append(entry)
            
            if is_rejected:
                print(f"\n   🚨 FINAL: REJECTED at {rejected_level}")
                if rejection_reason:
                    print(f"      Reason: {rejection_reason[:50]}...")
            else:
                print(f"\n   ✅ FINAL: {overall_status.upper()}")
        
        print(f"\n{'='*60}")
        print(f"✅ Returning {len(status_entries)} status entries")
        print(f"{'='*60}\n")
        
        return {"status_entries": status_entries}
        
    except Exception as e:
        print(f"\n❌❌ ERROR fetching status:")
        print(f"   {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
        
# HR API

@app.post("/api/ope/hr/approve/{employee_code}")
async def hr_approve_employee(
    employee_code: str,
    request: Request,
    current_user=Depends(get_current_user)
):
    try:
        hr_emp_code = current_user["employee_code"].strip().upper()
        employee_code = employee_code.strip().upper()
        
        body = await request.json()
        approval_remark = body.get("remark", "Approved without remark")
        
        print(f"\n{'='*60}")
        print(f"✅ HR APPROVAL REQUEST")
        print(f"HR: {hr_emp_code}")
        print(f"Employee: {employee_code}")
        print(f"Remark: {approval_remark}")
        print(f"{'='*60}\n")
        
        if hr_emp_code != "JHS729":
            raise HTTPException(status_code=403, detail="Only HR can perform this action")
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_code})
        
        if not ope_doc:
            raise HTTPException(status_code=404, detail="No OPE data found")
        
        approved_count = 0
        current_time = datetime.utcnow().isoformat()
        
        data_array = ope_doc.get("Data", [])
        payroll_months_approved = set()
        
        for i, data_item in enumerate(data_array):
            for month_range, entries in data_item.items():
                for j, entry in enumerate(entries):
                    entry_status = entry.get("status", "").lower()
                    
                    if entry_status in ["approved", "rejected"]:
                        await db["OPE_data"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"Data.{i}.{month_range}.{j}.status": "approved",
                                f"Data.{i}.{month_range}.{j}.hr_approved": True,
                                f"Data.{i}.{month_range}.{j}.hr_approved_by": hr_emp_code,
                                f"Data.{i}.{month_range}.{j}.hr_approved_date": current_time,
                                f"Data.{i}.{month_range}.{j}.approved_by": hr_emp_code,
                                f"Data.{i}.{month_range}.{j}.approver_name": "HR",
                                f"Data.{i}.{month_range}.{j}.approved_date": current_time,
                                f"Data.{i}.{month_range}.{j}.approval_remark": approval_remark,
                                f"Data.{i}.{month_range}.{j}.rejected_by": None,
                                f"Data.{i}.{month_range}.{j}.rejector_name": None,
                                f"Data.{i}.{month_range}.{j}.rejected_date": None,
                                f"Data.{i}.{month_range}.{j}.rejection_reason": None,
                                f"Data.{i}.{month_range}.{j}.rejected_level": None
                            }}
                        )
                        
                        payroll_months_approved.add(month_range)
                        approved_count += 1
        
        if approved_count == 0:
            raise HTTPException(status_code=404, detail="No entries found for HR approval")
        
        status_doc = await db["Status"].find_one({"employeeId": employee_code})
        
        if status_doc:
            approval_status = status_doc.get("approval_status", [])
            
            for i, ps in enumerate(approval_status):
                if ps.get("payroll_month") in payroll_months_approved:
                    total_levels = ps.get("total_levels", 2)
                    
                    if total_levels == 2:
                        await db["Status"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"approval_status.{i}.L2.status": True,
                                f"approval_status.{i}.L2.approver_code": hr_emp_code,
                                f"approval_status.{i}.L2.approver_name": "HR",
                                f"approval_status.{i}.L2.approved_date": current_time,
                                f"approval_status.{i}.L2.approval_remark": approval_remark,
                                f"approval_status.{i}.overall_status": "approved",
                                f"approval_status.{i}.current_level": "Completed"
                            }}
                        )
                    elif total_levels == 3:
                        await db["Status"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"approval_status.{i}.L3.status": True,
                                f"approval_status.{i}.L3.approver_code": hr_emp_code,
                                f"approval_status.{i}.L3.approver_name": "HR",
                                f"approval_status.{i}.L3.approved_date": current_time,
                                f"approval_status.{i}.L3.approval_remark": approval_remark,
                                f"approval_status.{i}.overall_status": "approved",
                                f"approval_status.{i}.current_level": "Completed"
                            }}
                        )
        
        hr_approved_doc = await db["HR_Approved"].find_one({"HR_Code": hr_emp_code})
        
        if not hr_approved_doc:
            await db["HR_Approved"].insert_one({
                "HR_Code": hr_emp_code,
                "EmployeesCodes": [employee_code]
            })
        else:
            if employee_code not in hr_approved_doc.get("EmployeesCodes", []):
                await db["HR_Approved"].update_one(
                    {"HR_Code": hr_emp_code},
                    {"$addToSet": {"EmployeesCodes": employee_code}}
                )
        
        await db["HR_Rejected"].update_one(
            {"HR_Code": hr_emp_code},
            {"$pull": {"EmployeesCodes": employee_code}}
        )
        
        print(f"✅ HR approved {approved_count} entries")
        print(f"   Approval Remark: {approval_remark}")
        
        return {
            "message": f"HR approved {approved_count} entries",
            "approved_count": approved_count,
            "approval_remark": approval_remark
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))        

@app.post("/api/ope/hr/reject/{employee_code}")
async def hr_reject_employee(
    employee_code: str,
    request: Request,
    current_user=Depends(get_current_user)
):
    try:
        hr_emp_code = current_user["employee_code"].strip().upper()
        employee_code = employee_code.strip().upper()
        
        body = await request.json()
        rejection_reason = body.get("reason", "No reason provided")
        
        print(f"\n{'='*60}")
        print(f"❌ HR REJECTION REQUEST")
        print(f"HR: {hr_emp_code}")
        print(f"Employee: {employee_code}")
        print(f"{'='*60}\n")
        
        if hr_emp_code != "JHS729":
            raise HTTPException(status_code=403, detail="Only HR can perform this action")
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_code})
        
        if not ope_doc:
            raise HTTPException(status_code=404, detail="No OPE data found")
        
        rejected_count = 0
        current_time = datetime.utcnow().isoformat()
        
        data_array = ope_doc.get("Data", [])
        payroll_months_rejected = set()
        
        for i, data_item in enumerate(data_array):
            for month_range, entries in data_item.items():
                for j, entry in enumerate(entries):
                    entry_status = entry.get("status", "").lower()
                    
                    if entry_status == "approved":
                        await db["OPE_data"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"Data.{i}.{month_range}.{j}.status": "rejected",
                                f"Data.{i}.{month_range}.{j}.rejected_by": hr_emp_code,
                                f"Data.{i}.{month_range}.{j}.rejector_name": "HR",
                                f"Data.{i}.{month_range}.{j}.rejected_date": current_time,
                                f"Data.{i}.{month_range}.{j}.rejection_reason": rejection_reason,
                                f"Data.{i}.{month_range}.{j}.rejected_level": "L2",
                                f"Data.{i}.{month_range}.{j}.hr_approved": False,
                                f"Data.{i}.{month_range}.{j}.hr_approved_by": None,
                                f"Data.{i}.{month_range}.{j}.hr_approved_date": None
                            }}
                        )
                        
                        payroll_months_rejected.add(month_range)
                        rejected_count += 1
        
        if rejected_count == 0:
            raise HTTPException(status_code=404, detail="No entries found for HR rejection")
        
        status_doc = await db["Status"].find_one({"employeeId": employee_code})
        
        if status_doc:
            approval_status = status_doc.get("approval_status", [])
            
            for i, ps in enumerate(approval_status):
                if ps.get("payroll_month") in payroll_months_rejected:
                    total_levels = ps.get("total_levels", 2)
                    
                    if total_levels == 2:
                        await db["Status"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"approval_status.{i}.L2.status": False,
                                f"approval_status.{i}.L2.rejected_by": hr_emp_code,
                                f"approval_status.{i}.L2.rejected_date": current_time,
                                f"approval_status.{i}.overall_status": "rejected",
                                f"approval_status.{i}.rejection_reason": rejection_reason
                            }}
                        )
                    elif total_levels == 3:
                        await db["Status"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"approval_status.{i}.L3.status": False,
                                f"approval_status.{i}.L3.rejected_by": hr_emp_code,
                                f"approval_status.{i}.L3.rejected_date": current_time,
                                f"approval_status.{i}.overall_status": "rejected",
                                f"approval_status.{i}.rejection_reason": rejection_reason
                            }}
                        )
        
        hr_rejected_doc = await db["HR_Rejected"].find_one({"HR_Code": hr_emp_code})
        
        if not hr_rejected_doc:
            await db["HR_Rejected"].insert_one({
                "HR_Code": hr_emp_code,
                "EmployeesCodes": [employee_code]
            })
        else:
            if employee_code not in hr_rejected_doc.get("EmployeesCodes", []):
                await db["HR_Rejected"].update_one(
                    {"HR_Code": hr_emp_code},
                    {"$addToSet": {"EmployeesCodes": employee_code}}
                )
        
        await db["HR_Approved"].update_one(
            {"HR_Code": hr_emp_code},
            {"$pull": {"EmployeesCodes": employee_code}}
        )
        
        print(f"✅ HR rejected {rejected_count} entries")
        
        return {
            "message": f"HR rejected {rejected_count} entries",
            "rejected_count": rejected_count
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    

@app.post("/api/ope/hr/is-hr")
async def check_if_hr(current_user=Depends(get_current_user)):
    try:
        emp_code = current_user["employee_code"].strip().upper()
        is_hr = (emp_code == "JHS729")
        
        return {
            "employee_code": emp_code,
            "is_hr": is_hr
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
@app.get("/api/ope/hr/approved-employees")
async def get_hr_approved_employees(current_user=Depends(get_current_user)):
    try:
        hr_emp_code = current_user["employee_code"].strip().upper()
        
        if hr_emp_code != "JHS729":
            raise HTTPException(status_code=403, detail="Only HR can access this")
        
        print(f"📋 Fetching HR approved employees")
        
        hr_approved_doc = await db["HR_Approved"].find_one({"HR_Code": hr_emp_code})
        
        employee_codes = []
        if hr_approved_doc:
            employee_codes = hr_approved_doc.get("EmployeesCodes", [])
        
        print(f"✅ Found {len(employee_codes)} HR approved employees")
        
        return {
            "employee_codes": employee_codes,
            "count": len(employee_codes)
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/ope/hr/rejected-employees")
async def get_hr_rejected_employees(current_user=Depends(get_current_user)):
    try:
        hr_emp_code = current_user["employee_code"].strip().upper()
        
        if hr_emp_code != "JHS729":
            raise HTTPException(status_code=403, detail="Only HR can access this")
        
        print(f"📋 Fetching HR rejected employees")
        
        hr_rejected_doc = await db["HR_Rejected"].find_one({"HR_Code": hr_emp_code})
        
        employee_codes = []
        if hr_rejected_doc:
            employee_codes = hr_rejected_doc.get("EmployeesCodes", [])
        
        print(f"✅ Found {len(employee_codes)} HR rejected employees")
        
        return {
            "employee_codes": employee_codes,
            "count": len(employee_codes)
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
        
@app.get("/api/check-partner/{employee_code}")
async def check_if_partner(employee_code: str, current_user=Depends(get_current_user)):
    try:
        emp_code = employee_code.strip().upper()
        
        print(f"🔍 Checking if {emp_code} is a Partner...")
        
        if current_user["employee_code"].upper() != emp_code:
            raise HTTPException(status_code=403, detail="Access denied")
        
        partner = await db["Partner"].find_one({"PartnerEmpCode": emp_code})
        
        is_partner = partner is not None
        
        if is_partner:
            print(f"✅ {emp_code} IS a Partner")
        else:
            print(f"❌ {emp_code} is NOT a Partner")
        
        return {
            "employee_code": emp_code,
            "isPartner": is_partner,
            "partner_name": partner.get("Partner_Name") if partner else None
        }
        
    except Exception as e:
        print(f"❌ Error checking partner role: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    

@app.get("/api/ope/partner/pending")
async def get_partner_pending(current_user: dict = Depends(get_current_user)):
    try:
        partner_code = current_user.get("employee_code")
        
        print(f"\n{'='*60}")
        print(f"🔍 PARTNER PENDING REQUEST FROM: {partner_code}")
        print(f"{'='*60}\n")
        
        partner_pending = await db["Pending"].find_one({"ReportingEmpCode": partner_code})
        
        if not partner_pending or not partner_pending.get("EmployeesCodes"):
            print("📭 No pending employees")
            return {"employees": []}
        
        pending_emp_codes = partner_pending.get("EmployeesCodes", [])
        print(f"📊 Found {len(pending_emp_codes)} employees in Pending")
        print(f"   Codes: {pending_emp_codes}")
        
        employees_list = []
        
        for emp_code in pending_emp_codes:
            print(f"\n{'='*40}")
            print(f"✅ Processing: {emp_code}")
            print(f"{'='*40}")
            
            status_docs = await db["Status"].find({
                "employeeId": emp_code
            }).to_list(length=None)
            
            print(f"   📋 Found {len(status_docs)} Status docs")
            
            if not status_docs:
                print(f"   ❌ No Status - SKIP")
                continue
            
            employee = await db["Employee"].find_one({"EmployeeId": emp_code})
            
            if not employee:
                print(f"   ⚠️ Employee not found - using fallback")
                employee = {
                    "EmployeeId": emp_code,
                    "EmployeeName": status_docs[0].get("employeeName", emp_code),
                    "Designation": "N/A",
                    "Department": "N/A",
                    "OPE_limit": 5000
                }
            
            print(f"   👤 Name: {employee.get('EmployeeName')}")
            
            ope_doc = await db["OPE_data"].find_one({"employeeId": emp_code})
            
            if not ope_doc:
                print(f"   ❌ No OPE_data document - SKIP")
                continue
            
            print(f"   ✅ OPE_data document found")
            
            for status_doc in status_docs:
                approval_status_raw = status_doc.get("approval_status")
                
                print(f"\n   📄 Status Document")
                print(f"      Type: {type(approval_status_raw).__name__}")
                
                if isinstance(approval_status_raw, list):
                    approval_status_array = approval_status_raw
                elif isinstance(approval_status_raw, dict):
                    approval_status_array = [approval_status_raw]
                else:
                    print(f"      ❌ Unknown format - SKIP")
                    continue
                
                for idx, approval_status in enumerate(approval_status_array):
                    print(f"\n   📋 Processing approval_status[{idx}]")
                    
                    month_range = approval_status.get("payroll_month") or approval_status.get("month_range")
                    total_levels = approval_status.get("total_levels", 2)
                    current_level = approval_status.get("current_level", "L1")
                    overall_status = approval_status.get("overall_status", "pending")
                    submitter_type = approval_status.get("submitter_type", "Employee")
                    
                    print(f"      Month: {month_range}")
                    print(f"      Total Levels: {total_levels}")
                    print(f"      Current Level: {current_level}")
                    print(f"      Overall Status: {overall_status}")
                    
                    if overall_status != "pending":
                        print(f"      ⚠️ Not pending - SKIP")
                        continue
                    
                    if not month_range:
                        print(f"      ⚠️ No month_range - SKIP")
                        continue
                    
                    partner_is_approver = False
                    l1_approver = "N/A"
                    l1_approved_date = None
                    
                    if current_level == "L1":
                        L1 = approval_status.get("L1", {})
                        l1_approver_code = L1.get("approver_code")
                        
                        print(f"      L1 Approver: {l1_approver_code}")
                        
                        if l1_approver_code == partner_code:
                            partner_is_approver = True
                            l1_approver = "Self (RM)" if submitter_type == "Reporting_Manager" else "Pending"
                            print(f"      ✅ MATCH! Partner is L1 approver")
                    
                    elif current_level == "L2":
                        L2 = approval_status.get("L2", {})
                        l2_approver_code = L2.get("approver_code")
                        
                        print(f"      L2 Approver: {l2_approver_code}")
                        
                        if l2_approver_code == partner_code:
                            partner_is_approver = True
                            L1 = approval_status.get("L1", {})
                            l1_approver = L1.get("approver_name", "Unknown")
                            l1_approved_date = L1.get("approved_date")
                            print(f"      ✅ MATCH! Partner is L2 approver")
                    
                    if not partner_is_approver:
                        print(f"      ❌ Partner NOT approver - SKIP")
                        continue
                    
                    print(f"      🔍 Looking for entries in nested Data structure")
                    
                    entries = []
                    data_array = ope_doc.get("Data", [])
                    
                    print(f"      📦 OPE_data.Data has {len(data_array)} items")
                    
                    for data_item in data_array:
                        if month_range in data_item:
                            month_entries = data_item[month_range]
                            print(f"      ✅ Found month '{month_range}' with {len(month_entries)} entries")
                            
                            entry_status = "pending" if submitter_type == "Reporting_Manager" else "approved"
                            
                            for entry in month_entries:
                                e_status = entry.get("status", "").lower()
                                
                                if e_status == entry_status or entry_status == "approved":
                                    entries.append(entry)
                                    print(f"         ✅ Entry: {entry.get('date')} | ₹{entry.get('amount')} | status={e_status}")
                            
                            break
                    
                    print(f"      📦 Total entries found: {len(entries)}")
                    
                    if not entries:
                        print(f"      ❌ No entries - SKIP")
                        continue
                    
                    total_amount = sum(float(e.get("amount", 0)) for e in entries)
                    
                    formatted_entries = []
                    for entry in entries:
                        formatted_entries.append({
                            "_id": str(entry.get("_id", "")),
                            "date": entry.get("date"),
                            "client": entry.get("client"),
                            "project_id": entry.get("project_id"),
                            "project_name": entry.get("project_name"),
                            "project_type": entry.get("project_type"),
                            "location_from": entry.get("location_from"),
                            "location_to": entry.get("location_to"),
                            "travel_mode": entry.get("travel_mode"),
                            "amount": entry.get("amount"),
                            "remarks": entry.get("remarks"),
                            "ticket_pdf": entry.get("ticket_pdf"),    # GridFS ID
                            "month_range": month_range,
                            "total_levels": total_levels,
                            "current_level": current_level,
                            "submitter_type": submitter_type
                        })
                    
                    employees_list.append({
                        "employeeId": emp_code,
                        "employeeName": employee.get("EmployeeName"),
                        "designation": employee.get("Designation", "N/A"),
                        "department": employee.get("Department", "N/A"),
                        "reportingManager": employee.get("Reporting_Manager", "N/A"),
                        "payroll_month": month_range,
                        "total_amount": total_amount,
                        "limit": employee.get("OPE_limit", 5000),
                        "pending_entries": len(entries),
                        "total_levels": total_levels,
                        "current_level": current_level,
                        "L1_approver": l1_approver,
                        "L1_approved_date": l1_approved_date,
                        "submission_date": approval_status.get("submission_date"),
                        "submitter_type": submitter_type,
                        "entries": formatted_entries
                    })
                    
                    print(f"      ✅ Added to response")
                    print(f"         Entries: {len(entries)}, Amount: ₹{total_amount}")
        
        print(f"\n{'='*60}")
        print(f"✅ FINAL: Returning {len(employees_list)} employees")
        print(f"{'='*60}\n")
        
        return {"employees": employees_list}
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ope/partner/approve/{employee_code}")
async def partner_approve_employee(
    employee_code: str,
    request: Request,
    current_user=Depends(get_current_user)
):
    try:
        partner_emp_code = current_user["employee_code"].strip().upper()
        employee_code = employee_code.strip().upper()
        
        body = await request.json()
        approval_remark = body.get("remark", "Approved without remark")
        
        print(f"\n{'='*60}")
        print(f"✅ PARTNER APPROVAL REQUEST")
        print(f"Partner: {partner_emp_code}")
        print(f"Employee: {employee_code}")
        print(f"Remark: {approval_remark}")
        print(f"{'='*60}\n")
        
        partner = await db["Partner"].find_one({"PartnerEmpCode": partner_emp_code})
        if not partner:
            raise HTTPException(status_code=403, detail="You are not a Partner")
        
        partner_name = partner.get("Partner_Name", partner_emp_code)
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_code})
        
        if not ope_doc:
            raise HTTPException(status_code=404, detail="No OPE data found")
        
        approved_count = 0
        current_time = datetime.utcnow().isoformat()
        
        data_array = ope_doc.get("Data", [])
        payroll_months_approved = set()
        
        print(f"📦 OPE_data.Data has {len(data_array)} items")
        
        for i, data_item in enumerate(data_array):
            for month_range, entries in data_item.items():
                print(f"\n📅 Processing month: {month_range}")
                print(f"   Entries: {len(entries)}")
                
                for j, entry in enumerate(entries):
                    entry_status = entry.get("status", "").lower()
                    
                    if entry_status in ["pending", "approved"]:
                        await db["OPE_data"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"Data.{i}.{month_range}.{j}.status": "approved",
                                f"Data.{i}.{month_range}.{j}.partner_approved": True,
                                f"Data.{i}.{month_range}.{j}.partner_approved_by": partner_emp_code,
                                f"Data.{i}.{month_range}.{j}.partner_approved_date": current_time,
                                f"Data.{i}.{month_range}.{j}.partner_name": partner_name,
                                f"Data.{i}.{month_range}.{j}.approval_remark": approval_remark,
                                f"Data.{i}.{month_range}.{j}.L2_approved": True,
                                f"Data.{i}.{month_range}.{j}.L2_approver_code": partner_emp_code,
                                f"Data.{i}.{month_range}.{j}.L2_approver_name": partner_name
                            }}
                        )
                        
                        payroll_months_approved.add(month_range)
                        approved_count += 1
                        print(f"   ✅ Approved entry {j+1}: {entry.get('date')} | ₹{entry.get('amount')}")
        
        if approved_count == 0:
            raise HTTPException(status_code=404, detail="No pending entries found for approval")
        
        print(f"\n✅ Total entries approved: {approved_count}")
        
        status_doc = await db["Status"].find_one({"employeeId": employee_code})
        
        if status_doc:
            approval_status_array = status_doc.get("approval_status", [])
            
            print(f"\n📊 Updating Status collection...")
            
            if isinstance(approval_status_array, dict):
                approval_status_array = [approval_status_array]
            
            for i, approval_status in enumerate(approval_status_array):
                payroll_month = approval_status.get("payroll_month") or approval_status.get("month_range")
                
                if payroll_month in payroll_months_approved:
                    total_levels = approval_status.get("total_levels", 2)
                    submitter_type = approval_status.get("submitter_type", "Employee")
                    
                    print(f"   📋 Month: {payroll_month}")
                    print(f"      Total Levels: {total_levels}")
                    print(f"      Submitter: {submitter_type}")
                    
                    if submitter_type == "Reporting_Manager":
                        await db["Status"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"approval_status.{i}.L1.status": True,
                                f"approval_status.{i}.L1.approver_code": partner_emp_code,
                                f"approval_status.{i}.L1.approver_name": partner_name,
                                f"approval_status.{i}.L1.approved_date": current_time,
                                f"approval_status.{i}.L1.approval_remark": approval_remark,
                                f"approval_status.{i}.current_level": "L2",
                                f"approval_status.{i}.overall_status": "pending"
                            }}
                        )
                        print(f"      ✅ RM: L1 approved → L2 (HR) pending")
                    
                    elif total_levels == 3:
                        await db["Status"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"approval_status.{i}.L2.status": True,
                                f"approval_status.{i}.L2.approver_code": partner_emp_code,
                                f"approval_status.{i}.L2.approver_name": partner_name,
                                f"approval_status.{i}.L2.approved_date": current_time,
                                f"approval_status.{i}.L2.approval_remark": approval_remark,
                                f"approval_status.{i}.current_level": "L3",
                                f"approval_status.{i}.overall_status": "pending"
                            }}
                        )
                        print(f"      ✅ 3-level: L2 approved → L3 (HR) pending")
                    
                    elif total_levels == 2:
                        await db["Status"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"approval_status.{i}.L1.status": True,
                                f"approval_status.{i}.L1.approver_code": partner_emp_code,
                                f"approval_status.{i}.L1.approver_name": partner_name,
                                f"approval_status.{i}.L1.approved_date": current_time,
                                f"approval_status.{i}.L1.approval_remark": approval_remark,
                                f"approval_status.{i}.current_level": "L2",
                                f"approval_status.{i}.overall_status": "pending"
                            }}
                        )
                        print(f"      ✅ 2-level: L1 approved → L2 (HR) pending")
        
        await db["Pending"].update_one(
            {"ReportingEmpCode": partner_emp_code},
            {"$pull": {"EmployeesCodes": employee_code}}
        )
        print(f"\n✅ Removed from Partner's Pending")
        
        partner_approved_doc = await db["Partner_Approved"].find_one(
            {"PartnerEmpCode": partner_emp_code}
        )
        
        if not partner_approved_doc:
            await db["Partner_Approved"].insert_one({
                "PartnerEmpCode": partner_emp_code,
                "EmployeesCodes": [employee_code]
            })
            print(f"✅ Created NEW Partner_Approved document")
        else:
            if employee_code not in partner_approved_doc.get("EmployeesCodes", []):
                await db["Partner_Approved"].update_one(
                    {"PartnerEmpCode": partner_emp_code},
                    {"$addToSet": {"EmployeesCodes": employee_code}}
                )
                print(f"✅ Added to Partner_Approved collection")
        
        print(f"\n{'='*60}")
        print(f"✅ Partner approved {approved_count} entries for {employee_code}")
        print(f"   Approval Remark: {approval_remark}")
        print(f"   Next Level: HR")
        print(f"{'='*60}\n")
        
        return {
            "message": f"Successfully approved {approved_count} entries",
            "approved_count": approved_count,
            "approval_remark": approval_remark,
            "next_level": "HR"
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error in partner approve: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    

@app.get("/api/debug/partner-pending/{partner_code}")
async def debug_partner_pending(partner_code: str):
    partner_code = partner_code.strip().upper()
    
    pending = await db["Pending"].find_one({"ReportingEmpCode": partner_code})
    emp_codes = pending.get("EmployeesCodes", []) if pending else []
    
    result = {"pending_employees": emp_codes, "status_details": []}
    
    for emp_code in emp_codes:
        status_docs = await db["Status"].find({"employeeId": emp_code}).to_list(length=None)
        
        for sdoc in status_docs:
            approval_arr = sdoc.get("approval_status", [])
            if isinstance(approval_arr, dict):
                approval_arr = [approval_arr]
            
            for ps in approval_arr:
                result["status_details"].append({
                    "employeeId": emp_code,
                    "payroll_month": ps.get("payroll_month"),
                    "current_level": ps.get("current_level"),
                    "overall_status": ps.get("overall_status"),
                    "submitter_type": ps.get("submitter_type"),
                    "L1_approver_code": ps.get("L1", {}).get("approver_code"),
                    "L2_approver_code": ps.get("L2", {}).get("approver_code"),
                    "partner_code_being_checked": partner_code
                })
    
    return result


@app.post("/api/ope/partner/reject/{employee_code}")
async def partner_reject_employee(
    employee_code: str,
    request: Request,
    current_user=Depends(get_current_user)
):
    try:
        partner_emp_code = current_user["employee_code"].strip().upper()
        employee_code = employee_code.strip().upper()
        
        body = await request.json()
        rejection_reason = body.get("reason", "No reason provided")
        
        print(f"\n{'='*60}")
        print(f"❌ PARTNER REJECTION REQUEST")
        print(f"Partner: {partner_emp_code}")
        print(f"Employee: {employee_code}")
        print(f"Reason: {rejection_reason}")
        print(f"{'='*60}\n")
        
        partner = await db["Partner"].find_one({"PartnerEmpCode": partner_emp_code})
        if not partner:
            raise HTTPException(status_code=403, detail="You are not a Partner")
        
        partner_name = partner.get("Partner_Name", partner_emp_code)
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_code})
        
        if not ope_doc:
            raise HTTPException(status_code=404, detail="No OPE data found")
        
        rejected_count = 0
        current_time = datetime.utcnow().isoformat()
        
        data_array = ope_doc.get("Data", [])
        payroll_months_rejected = set()
        
        print(f"📦 OPE_data.Data has {len(data_array)} items")
        
        for i, data_item in enumerate(data_array):
            for month_range, entries in data_item.items():
                print(f"\n📅 Processing month: {month_range}")
                print(f"   Entries: {len(entries)}")
                
                for j, entry in enumerate(entries):
                    entry_status = entry.get("status", "").lower()
                    
                    if entry_status in ["pending", "approved"]:
                        await db["OPE_data"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"Data.{i}.{month_range}.{j}.status": "rejected",
                                f"Data.{i}.{month_range}.{j}.rejected_by": partner_emp_code,
                                f"Data.{i}.{month_range}.{j}.rejector_name": partner_name,
                                f"Data.{i}.{month_range}.{j}.rejected_date": current_time,
                                f"Data.{i}.{month_range}.{j}.rejection_reason": rejection_reason,
                                f"Data.{i}.{month_range}.{j}.rejected_level": "L2"
                            }}
                        )
                        
                        payroll_months_rejected.add(month_range)
                        rejected_count += 1
                        print(f"   ❌ Rejected entry {j+1}: {entry.get('date')} | ₹{entry.get('amount')}")
        
        if rejected_count == 0:
            raise HTTPException(status_code=404, detail="No pending entries found for rejection")
        
        print(f"\n❌ Total entries rejected: {rejected_count}")
        
        status_doc = await db["Status"].find_one({"employeeId": employee_code})
        
        if status_doc:
            approval_status_array = status_doc.get("approval_status", [])
            
            print(f"\n📊 Updating Status collection...")
            
            if isinstance(approval_status_array, dict):
                approval_status_array = [approval_status_array]
            
            for i, approval_status in enumerate(approval_status_array):
                payroll_month = approval_status.get("payroll_month") or approval_status.get("month_range")
                
                if payroll_month in payroll_months_rejected:
                    total_levels = approval_status.get("total_levels", 2)
                    submitter_type = approval_status.get("submitter_type", "Employee")
                    
                    print(f"   📋 Month: {payroll_month}")
                    print(f"      Total Levels: {total_levels}")
                    print(f"      Submitter: {submitter_type}")
                    
                    if submitter_type == "Reporting_Manager":
                        await db["Status"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"approval_status.{i}.L1.status": False,
                                f"approval_status.{i}.L1.rejected_by": partner_emp_code,
                                f"approval_status.{i}.L1.rejected_date": current_time,
                                f"approval_status.{i}.overall_status": "rejected",
                                f"approval_status.{i}.rejection_reason": rejection_reason,
                                f"approval_status.{i}.rejected_level": "L1"
                            }}
                        )
                        print(f"      ❌ RM: L1 rejected")
                    
                    elif total_levels == 3:
                        await db["Status"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"approval_status.{i}.L2.status": False,
                                f"approval_status.{i}.L2.rejected_by": partner_emp_code,
                                f"approval_status.{i}.L2.rejected_date": current_time,
                                f"approval_status.{i}.overall_status": "rejected",
                                f"approval_status.{i}.rejection_reason": rejection_reason,
                                f"approval_status.{i}.rejected_level": "L2"
                            }}
                        )
                        print(f"      ❌ 3-level: L2 rejected")
                    
                    elif total_levels == 2:
                        await db["Status"].update_one(
                            {"employeeId": employee_code},
                            {"$set": {
                                f"approval_status.{i}.L1.status": False,
                                f"approval_status.{i}.L1.rejected_by": partner_emp_code,
                                f"approval_status.{i}.L1.rejected_date": current_time,
                                f"approval_status.{i}.overall_status": "rejected",
                                f"approval_status.{i}.rejection_reason": rejection_reason,
                                f"approval_status.{i}.rejected_level": "L1"
                            }}
                        )
                        print(f"      ❌ 2-level: L1 rejected")
        
        await db["Pending"].update_one(
            {"ReportingEmpCode": partner_emp_code},
            {"$pull": {"EmployeesCodes": employee_code}}
        )
        print(f"\n✅ Removed from Partner's Pending")
        
        partner_rejected_doc = await db["Partner_Rejected"].find_one(
            {"PartnerEmpCode": partner_emp_code}
        )
        
        if not partner_rejected_doc:
            await db["Partner_Rejected"].insert_one({
                "PartnerEmpCode": partner_emp_code,
                "EmployeesCodes": [employee_code]
            })
            print(f"✅ Created NEW Partner_Rejected document")
        else:
            if employee_code not in partner_rejected_doc.get("EmployeesCodes", []):
                await db["Partner_Rejected"].update_one(
                    {"PartnerEmpCode": partner_emp_code},
                    {"$addToSet": {"EmployeesCodes": employee_code}}
                )
                print(f"✅ Added to Partner_Rejected collection")
        
        print(f"\n{'='*60}")
        print(f"❌ Partner rejected {rejected_count} entries for {employee_code}")
        print(f"{'='*60}\n")
        
        return {
            "message": f"Successfully rejected {rejected_count} entries",
            "rejected_count": rejected_count,
            "rejection_reason": rejection_reason
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error in partner reject: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/ope/partner/approved-list")
async def get_partner_approved_list(current_user=Depends(get_current_user)):
    try:
        partner_emp_code = current_user["employee_code"].strip().upper()
        
        partner = await db["Partner"].find_one({"PartnerEmpCode": partner_emp_code})
        if not partner:
            raise HTTPException(status_code=403, detail="You are not a Partner")
        
        partner_approved_doc = await db["Partner_Approved"].find_one(
            {"PartnerEmpCode": partner_emp_code}
        )
        
        employee_codes = []
        if partner_approved_doc:
            employee_codes = partner_approved_doc.get("EmployeesCodes", [])
        
        return {
            "partner_code": partner_emp_code,
            "employee_codes": employee_codes,
            "count": len(employee_codes)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/check-user-role/{employee_code}")
async def check_user_role_unified(employee_code: str, current_user=Depends(get_current_user)):
    try:
        emp_code = employee_code.strip().upper()
        
        if current_user["employee_code"].upper() != emp_code:
            raise HTTPException(status_code=403, detail="Access denied")
        
        print(f"🔍 Checking unified role for: {emp_code}")
        
        role_name = "Employee"
        is_hr = False
        is_partner = False
        is_manager = False
        is_employee = False
        has_approval_permissions = False
        additional_info = {}
        
        if emp_code == "JHS729":
            is_hr = True
            role_name = "HR"
            has_approval_permissions = True
            print(f"👔 {emp_code} is HR")
            
        elif await db["Partner"].find_one({"PartnerEmpCode": emp_code}):
            partner = await db["Partner"].find_one({"PartnerEmpCode": emp_code})
            is_partner = True
            role_name = "Partner"
            has_approval_permissions = True
            additional_info["partner_name"] = partner.get("Partner_Name")
            print(f"👔 {emp_code} is a Partner")
            
        elif await db["Reporting_managers"].find_one({"ReportingEmpCode": emp_code}):
            manager = await db["Reporting_managers"].find_one({"ReportingEmpCode": emp_code})
            is_manager = True
            role_name = "Reporting Manager"
            has_approval_permissions = True
            additional_info["manager_name"] = manager.get("ReportingEmpName")
            additional_info["email"] = manager.get("Email ID")
            print(f"👔 {emp_code} is a Reporting Manager")
            
        else:
            is_employee = True
            role_name = "Employee"
            has_approval_permissions = False
            print(f"👤 {emp_code} is a regular Employee")
        
        return {
            "employee_code": emp_code,
            "role": role_name,
            "is_hr": is_hr,
            "is_partner": is_partner,
            "is_manager": is_manager,
            "is_employee": is_employee,
            "has_approval_permissions": has_approval_permissions,
            "additional_info": additional_info
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error checking user role: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/ope/partner/rejected-list")
async def get_partner_rejected_list(current_user=Depends(get_current_user)):
    try:
        partner_emp_code = current_user["employee_code"].strip().upper()
        
        partner = await db["Partner"].find_one({"PartnerEmpCode": partner_emp_code})
        if not partner:
            raise HTTPException(status_code=403, detail="You are not a Partner")
        
        partner_rejected_doc = await db["Partner_Rejected"].find_one(
            {"PartnerEmpCode": partner_emp_code}
        )
        
        employee_codes = []
        if partner_rejected_doc:
            employee_codes = partner_rejected_doc.get("EmployeesCodes", [])
        
        return {
            "partner_code": partner_emp_code,
            "employee_codes": employee_codes,
            "count": len(employee_codes)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    

@app.get("/api/debug/check-partner-data/{employee_code}")
async def debug_check_partner_data(employee_code: str):
    try:
        partner_code = employee_code.strip().upper()
        
        status_docs = await db["Status"].find({}).to_list(length=None)
        
        partner_matches = []
        for doc in status_docs:
            emp_id = doc.get("employeeId")
            approval_statuses = doc.get("approval_status", [])
            
            for ps in approval_statuses:
                if ps.get("total_levels") == 3:
                    l2_code = ps.get("L2", {}).get("approver_code", "")
                    if l2_code == partner_code:
                        partner_matches.append({
                            "employeeId": emp_id,
                            "payroll_month": ps.get("payroll_month"),
                            "total_levels": ps.get("total_levels"),
                            "current_level": ps.get("current_level"),
                            "L1_status": ps.get("L1", {}).get("status"),
                            "L2_approver_code": l2_code,
                            "overall_status": ps.get("overall_status")
                        })
        
        return {
            "partner_code": partner_code,
            "total_status_docs": len(status_docs),
            "partner_matches": partner_matches,
            "match_count": len(partner_matches)
        }
        
    except Exception as e:
        return {"error": str(e)}


@app.post("/api/ope/partner/reject-single")
async def partner_reject_single_entry(
    request: Request,
    current_user=Depends(get_current_user)
):
    try:
        partner_emp_code = current_user["employee_code"].strip().upper()
        
        body = await request.json()
        entry_id = body.get("entry_id")
        employee_id = body.get("employee_id")
        reason = body.get("reason", "No reason provided")
        
        print(f"\n{'='*60}")
        print(f"❌ PARTNER REJECT SINGLE ENTRY")
        print(f"Partner: {partner_emp_code}")
        print(f"Employee: {employee_id}")
        print(f"Entry ID: {entry_id}")
        print(f"Reason: {reason}")
        print(f"{'='*60}\n")
        
        partner = await db["Partner"].find_one({"PartnerEmpCode": partner_emp_code})
        if not partner:
            raise HTTPException(status_code=403, detail="You are not a Partner")
        
        partner_name = partner.get("Partner_Name", partner_emp_code)
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_id})
        
        if not ope_doc:
            raise HTTPException(status_code=404, detail="Employee data not found")
        
        data_array = ope_doc.get("Data", [])
        updated = False
        current_time = datetime.utcnow().isoformat()
        payroll_month = None
        
        for i, data_item in enumerate(data_array):
            for month_range, entries in data_item.items():
                for j, entry in enumerate(entries):
                    if str(entry.get("_id")) == entry_id and entry.get("status") == "approved":
                        payroll_month = month_range
                        
                        print(f"✅ Found entry: Data.{i}.{month_range}.{j}")
                        
                        await db["OPE_data"].update_one(
                            {"employeeId": employee_id},
                            {"$set": {
                                f"Data.{i}.{month_range}.{j}.status": "rejected",
                                f"Data.{i}.{month_range}.{j}.rejected_by": partner_emp_code,
                                f"Data.{i}.{month_range}.{j}.rejector_name": partner_name,
                                f"Data.{i}.{month_range}.{j}.rejected_date": current_time,
                                f"Data.{i}.{month_range}.{j}.rejection_reason": reason,
                                f"Data.{i}.{month_range}.{j}.rejected_level": "L2",
                                f"Data.{i}.{month_range}.{j}.partner_approved": False,
                                f"Data.{i}.{month_range}.{j}.partner_approved_by": None,
                                f"Data.{i}.{month_range}.{j}.partner_approved_date": None
                            }}
                        )
                        
                        updated = True
                        print(f"✅ Entry rejected: {entry.get('date')} | ₹{entry.get('amount')}")
                        break
            if updated:
                break
        
        if not updated:
            raise HTTPException(status_code=404, detail="Entry not found or not approved")
        
        all_rejected = True
        any_approved = False
        
        for i, data_item in enumerate(data_array):
            for month_range, entries in data_item.items():
                for entry in entries:
                    if str(entry.get("_id")) != entry_id:
                        entry_status = entry.get("status", "").lower()
                        if entry_status == "approved":
                            any_approved = True
                            all_rejected = False
                            break
        
        print(f"\n📊 Status check:")
        print(f"   Any approved remaining: {any_approved}")
        print(f"   All rejected: {all_rejected}")
        
        if all_rejected and payroll_month:
            print(f"🔄 All entries rejected - updating Status")
            
            status_doc = await db["Status"].find_one({"employeeId": employee_id})
            
            if status_doc:
                approval_status_array = status_doc.get("approval_status", [])
                
                if isinstance(approval_status_array, dict):
                    approval_status_array = [approval_status_array]
                
                for i, approval_status in enumerate(approval_status_array):
                    pm = approval_status.get("payroll_month") or approval_status.get("month_range")
                    
                    if pm == payroll_month:
                        submitter_type = approval_status.get("submitter_type", "Employee")
                        total_levels = approval_status.get("total_levels", 2)
                        
                        if submitter_type == "Reporting_Manager":
                            level_key = "L1"
                        elif total_levels == 3:
                            level_key = "L2"
                        else:
                            level_key = "L1"
                        
                        await db["Status"].update_one(
                            {"employeeId": employee_id},
                            {"$set": {
                                f"approval_status.{i}.{level_key}.status": False,
                                f"approval_status.{i}.{level_key}.rejected_by": partner_emp_code,
                                f"approval_status.{i}.{level_key}.rejected_date": current_time,
                                f"approval_status.{i}.overall_status": "rejected",
                                f"approval_status.{i}.rejection_reason": reason,
                                f"approval_status.{i}.rejected_level": level_key
                            }}
                        )
                        print(f"✅ Status updated: {level_key} rejected")
                        break
        
        if all_rejected:
            print(f"🔄 Moving employee: Approved → Rejected")
            
            await db["Partner_Approved"].update_one(
                {"PartnerEmpCode": partner_emp_code},
                {"$pull": {"EmployeesCodes": employee_id}}
            )
            print(f"✅ Removed from Partner_Approved")
            
            partner_rejected_doc = await db["Partner_Rejected"].find_one(
                {"PartnerEmpCode": partner_emp_code}
            )
            
            if not partner_rejected_doc:
                await db["Partner_Rejected"].insert_one({
                    "PartnerEmpCode": partner_emp_code,
                    "EmployeesCodes": [employee_id]
                })
                print(f"✅ Created NEW Partner_Rejected document")
            else:
                if employee_id not in partner_rejected_doc.get("EmployeesCodes", []):
                    await db["Partner_Rejected"].update_one(
                        {"PartnerEmpCode": partner_emp_code},
                        {"$addToSet": {"EmployeesCodes": employee_id}}
                    )
                    print(f"✅ Added to Partner_Rejected")
        
        print(f"{'='*60}\n")
        
        return {
            "message": "Entry rejected successfully",
            "moved_to_rejected": all_rejected
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ope/partner/approve-single")
async def partner_approve_single_entry(
    request: Request,
    current_user=Depends(get_current_user)
):
    try:
        partner_emp_code = current_user["employee_code"].strip().upper()
        
        body = await request.json()
        entry_id = body.get("entry_id")
        employee_id = body.get("employee_id")
        
        print(f"\n{'='*60}")
        print(f"✅ PARTNER APPROVE SINGLE ENTRY")
        print(f"Partner: {partner_emp_code}")
        print(f"Employee: {employee_id}")
        print(f"Entry ID: {entry_id}")
        print(f"{'='*60}\n")
        
        partner = await db["Partner"].find_one({"PartnerEmpCode": partner_emp_code})
        if not partner:
            raise HTTPException(status_code=403, detail="You are not a Partner")
        
        partner_name = partner.get("Partner_Name", partner_emp_code)
        
        ope_doc = await db["OPE_data"].find_one({"employeeId": employee_id})
        
        if not ope_doc:
            raise HTTPException(status_code=404, detail="Employee data not found")
        
        data_array = ope_doc.get("Data", [])
        updated = False
        current_time = datetime.utcnow().isoformat()
        payroll_month = None
        
        for i, data_item in enumerate(data_array):
            for month_range, entries in data_item.items():
                for j, entry in enumerate(entries):
                    if str(entry.get("_id")) == entry_id and entry.get("status") == "rejected":
                        payroll_month = month_range
                        
                        print(f"✅ Found entry: Data.{i}.{month_range}.{j}")
                        
                        await db["OPE_data"].update_one(
                            {"employeeId": employee_id},
                            {"$set": {
                                f"Data.{i}.{month_range}.{j}.status": "approved",
                                f"Data.{i}.{month_range}.{j}.partner_approved": True,
                                f"Data.{i}.{month_range}.{j}.partner_approved_by": partner_emp_code,
                                f"Data.{i}.{month_range}.{j}.partner_approved_date": current_time,
                                f"Data.{i}.{month_range}.{j}.partner_name": partner_name,
                                f"Data.{i}.{month_range}.{j}.rejected_by": None,
                                f"Data.{i}.{month_range}.{j}.rejector_name": None,
                                f"Data.{i}.{month_range}.{j}.rejected_date": None,
                                f"Data.{i}.{month_range}.{j}.rejection_reason": None,
                                f"Data.{i}.{month_range}.{j}.rejected_level": None
                            }}
                        )
                        
                        updated = True
                        print(f"✅ Entry approved: {entry.get('date')} | ₹{entry.get('amount')}")
                        break
            if updated:
                break
        
        if not updated:
            raise HTTPException(status_code=404, detail="Entry not found or not rejected")
        
        no_rejected = True
        any_rejected = False
        
        for i, data_item in enumerate(data_array):
            for month_range, entries in data_item.items():
                for entry in entries:
                    if str(entry.get("_id")) != entry_id:
                        entry_status = entry.get("status", "").lower()
                        if entry_status == "rejected":
                            any_rejected = True
                            no_rejected = False
                            break
        
        print(f"\n📊 Status check:")
        print(f"   Any rejected remaining: {any_rejected}")
        print(f"   No rejected: {no_rejected}")
        
        if no_rejected and payroll_month:
            print(f"🔄 No rejected entries - updating Status")
            
            status_doc = await db["Status"].find_one({"employeeId": employee_id})
            
            if status_doc:
                approval_status_array = status_doc.get("approval_status", [])
                
                if isinstance(approval_status_array, dict):
                    approval_status_array = [approval_status_array]
                
                for i, approval_status in enumerate(approval_status_array):
                    pm = approval_status.get("payroll_month") or approval_status.get("month_range")
                    
                    if pm == payroll_month:
                        submitter_type = approval_status.get("submitter_type", "Employee")
                        total_levels = approval_status.get("total_levels", 2)
                        
                        if submitter_type == "Reporting_Manager":
                            level_key = "L1"
                            next_level = "L2"
                        elif total_levels == 3:
                            level_key = "L2"
                            next_level = "L3"
                        else:
                            level_key = "L1"
                            next_level = "L2"
                        
                        await db["Status"].update_one(
                            {"employeeId": employee_id},
                            {"$set": {
                                f"approval_status.{i}.{level_key}.status": True,
                                f"approval_status.{i}.{level_key}.approver_code": partner_emp_code,
                                f"approval_status.{i}.{level_key}.approver_name": partner_name,
                                f"approval_status.{i}.{level_key}.approved_date": current_time,
                                f"approval_status.{i}.{level_key}.rejected_by": None,
                                f"approval_status.{i}.{level_key}.rejected_date": None,
                                f"approval_status.{i}.overall_status": "pending",
                                f"approval_status.{i}.current_level": next_level,
                                f"approval_status.{i}.rejection_reason": None,
                                f"approval_status.{i}.rejected_level": None
                            }}
                        )
                        print(f"✅ Status updated: {level_key} approved → {next_level} pending")
                        break
        
        if no_rejected:
            print(f"🔄 Moving employee: Rejected → Approved")
            
            await db["Partner_Rejected"].update_one(
                {"PartnerEmpCode": partner_emp_code},
                {"$pull": {"EmployeesCodes": employee_id}}
            )
            print(f"✅ Removed from Partner_Rejected")
            
            partner_approved_doc = await db["Partner_Approved"].find_one(
                {"PartnerEmpCode": partner_emp_code}
            )
            
            if not partner_approved_doc:
                await db["Partner_Approved"].insert_one({
                    "PartnerEmpCode": partner_emp_code,
                    "EmployeesCodes": [employee_id]
                })
                print(f"✅ Created NEW Partner_Approved document")
            else:
                if employee_id not in partner_approved_doc.get("EmployeesCodes", []):
                    await db["Partner_Approved"].update_one(
                        {"PartnerEmpCode": partner_emp_code},
                        {"$addToSet": {"EmployeesCodes": employee_id}}
                    )
                    print(f"✅ Added to Partner_Approved")
        
        print(f"{'='*60}\n")
        
        return {
            "message": "Entry approved successfully",
            "moved_to_approved": no_rejected
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/check-admin/{employee_code}")
async def check_admin(employee_code: str, current_user=Depends(get_current_user)):
    """
    Check if an employee is admin
    Returns: { "isAdmin": true/false }
    """
    try:
        admin_collection = db["Admin"]
        admin_doc = await admin_collection.find_one({})
        print(admin_doc)
        
        if not admin_doc:
            return {"isAdmin": False}
        
        is_admin = employee_code.upper() in [code.upper() for code in admin_doc.get("employee_codes", [])]
        print(f"Is admin: {is_admin}")
        return {"isAdmin": is_admin}
        
    except Exception as e:
        print(f"Error checking admin: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================
# ADMIN APIs - Add these to the END of main.py (before the static mount)
# ============================================================

from fastapi import Query
from fastapi.responses import StreamingResponse
import io
from collections import defaultdict

# Try to import openpyxl; install if missing
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.run(["pip", "install", "openpyxl", "--break-system-packages"], capture_output=True)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ── Helper: verify admin ──────────────────────────────────────
async def verify_admin(current_user: dict):
    emp_code = current_user["employee_code"].strip().upper()
    admin_collection = db["Admin"]
    admin_doc = await admin_collection.find_one({})
    if not admin_doc:
        raise HTTPException(status_code=403, detail="Not authorized")
    codes = [c.upper() for c in admin_doc.get("employee_codes", [])]
    if emp_code not in codes:
        raise HTTPException(status_code=403, detail="Not authorized")
    return emp_code


# ── Helper: get all OPE entries (flat list) ───────────────────
async def get_all_ope_entries(payroll_month: str = None):
    """
    Returns list of dicts with employee + entry info flattened.
    Optionally filter by payroll_month.
    """
    entries = []
    status_docs = await db["Status"].find({}).to_list(length=None)

    # Build a lookup: employeeId -> list of payroll months
    emp_months = {}
    for sdoc in status_docs:
        eid = sdoc.get("employeeId")
        emp_name = sdoc.get("employeeName", "")
        for ps in sdoc.get("approval_status", []):
            pm = ps.get("payroll_month") or ps.get("month_range")
            if payroll_month and pm != payroll_month:
                continue
            emp_months.setdefault(eid, []).append({
                "payroll_month": pm,
                "total_amount": ps.get("total_amount", 0),
                "limit": ps.get("limit", 0),
                "overall_status": ps.get("overall_status", "pending"),
                "current_level": ps.get("current_level", "L1"),
                "total_levels": ps.get("total_levels", 2),
                "employee_name": emp_name,
            })

    # Fetch OPE_data for each employee
    all_ope = await db["OPE_data"].find({}).to_list(length=None)
    ope_map = {doc.get("employeeId"): doc for doc in all_ope}

    for emp_id, month_list in emp_months.items():
        ope_doc = ope_map.get(emp_id)
        for month_info in month_list:
            pm = month_info["payroll_month"]
            raw_entries = []
            if ope_doc:
                for data_item in ope_doc.get("Data", []):
                    if pm in data_item:
                        raw_entries = data_item[pm]
                        break

            for e in raw_entries:
                entries.append({
                    "employee_id": emp_id,
                    "employee_name": month_info["employee_name"] or ope_doc.get("employeeName", ""),
                    "payroll_month": pm,
                    "total_amount": month_info["total_amount"],
                    "limit": month_info["limit"],
                    "overall_status": month_info["overall_status"],
                    "current_level": month_info["current_level"],
                    "total_levels": month_info["total_levels"],
                    # entry fields
                    "date": e.get("date"),
                    "client": e.get("client", ""),
                    "project_id": e.get("project_id", ""),
                    "project_name": e.get("project_name", ""),
                    "project_type": e.get("project_type", ""),
                    "location_from": e.get("location_from", ""),
                    "location_to": e.get("location_to", ""),
                    "travel_mode": e.get("travel_mode", ""),
                    "amount": safe_float(e.get("amount", 0)),
                    "remarks": e.get("remarks", ""),
                    "status": e.get("status", ""),
                })
    return entries


# ── 1. ADMIN DASHBOARD ────────────────────────────────────────
@app.get("/api/admin/dashboard")
async def admin_dashboard(
    payroll_month: str = Query(None),
    emp_name: str = Query(None),
    emp_id: str = Query(None),
    current_user: dict = Depends(get_current_user)
):
    await verify_admin(current_user)

    # --- All status docs for KPI ---
    all_status = await db["Status"].find({}).to_list(length=None)

    # Build flat records for table & charts
    table_rows = []
    partner_totals = defaultdict(float)
    payroll_totals = defaultdict(float)
    all_months = set()
    total_employees_set = set()
    total_amount_all = 0.0
    amount_greater_count = 0

    # Employee details lookup for partner info
    emp_details_list = await db["Employee_details"].find({}).to_list(length=None)
    emp_partner_map = {e.get("EmpID", ""): e.get("Partner", "Unknown") for e in emp_details_list}

    for sdoc in all_status:
        emp_id_val = sdoc.get("employeeId", "")
        emp_name_val = sdoc.get("employeeName", "")
        partner = emp_partner_map.get(emp_id_val, "Unknown")

        for ps in sdoc.get("approval_status", []):
            pm = ps.get("payroll_month") or ps.get("month_range", "")
            if not pm:
                continue
            all_months.add(pm)

            total_amt = safe_float(ps.get("total_amount", 0))
            limit_val = safe_float(ps.get("limit", 0))
            ope_label = ps.get("ope_label", "")
            overall_st = ps.get("overall_status", "pending")
            curr_level = ps.get("current_level", "L1")

            # Apply filters
            if payroll_month and pm != payroll_month:
                continue
            if emp_name and emp_name.lower() not in emp_name_val.lower():
                continue
            if emp_id and emp_id.upper() not in emp_id_val.upper():
                continue

            total_employees_set.add(emp_id_val)
            total_amount_all += total_amt
            partner_totals[partner] += total_amt
            payroll_totals[pm] += total_amt

            if ope_label == "Greater" or (limit_val > 0 and total_amt > limit_val):
                amount_greater_count += 1

            table_rows.append({
                "employee_id": emp_id_val,
                "employee_name": emp_name_val,
                "payroll_month": pm,
                "total_amount": total_amt,
                "limit": limit_val,
                "overall_status": overall_st,
                "current_level": curr_level,
                "ope_label": ope_label,
            })

    # Payroll diff (compare last 2 months)
    sorted_months = sorted(payroll_totals.keys())
    payroll_diff = {}
    if len(sorted_months) >= 2:
        prev_month = sorted_months[-2]
        curr_month = sorted_months[-1]
        prev_total = payroll_totals[prev_month]
        curr_total = payroll_totals[curr_month]
        diff = abs(curr_total - prev_total)
        direction = "up" if curr_total > prev_total else ("down" if curr_total < prev_total else "same")
        payroll_diff = {
            "current_month": curr_month,
            "previous_month": prev_month,
            "difference": round(diff, 2),
            "direction": direction,
        }

    # Top 10 client chart — from OPE_data
    client_totals = defaultdict(float)
    all_ope_docs = await db["OPE_data"].find({}).to_list(length=None)
    for odoc in all_ope_docs:
        for data_item in odoc.get("Data", []):
            for pm_key, elist in data_item.items():
                if payroll_month and pm_key != payroll_month:
                    continue
                for e in elist:
                    client_totals[e.get("client", "Unknown")] += safe_float(e.get("amount", 0))

    top10_clients = sorted(client_totals.items(), key=lambda x: x[1], reverse=True)[:10]

    return {
        "kpis": {
            "total_employees": len(total_employees_set),
            "total_amount": round(total_amount_all, 2),
            "total_amount_greater": amount_greater_count,
            "payroll_diff": payroll_diff,
        },
        "charts": {
            "partner_wise": [{"_id": k, "total": round(v, 2)} for k, v in sorted(partner_totals.items(), key=lambda x: x[1], reverse=True)],
            "payroll_wise": [{"_id": k, "total": round(v, 2)} for k, v in sorted(payroll_totals.items())],
            "client_wise": [{"_id": k, "total": round(v, 2)} for k, v in top10_clients],
        },
        "table": table_rows,
        "all_payroll_months": sorted(all_months),
    }


# ── 2. CLIENT-WISE ANALYSIS ───────────────────────────────────
@app.get("/api/admin/client-analysis")
async def admin_client_analysis(
    payroll_month: str = Query(None),
    current_user: dict = Depends(get_current_user)
):
    """
    Per-client deep analysis:
    - How many employees claiming per client (payroll-month wise)
    - Amount difference between employees for same client (descending)
    - Flag: if one employee uses 'pass' and another uses 'ticket' for same client
    """
    await verify_admin(current_user)

    all_ope_docs = await db["OPE_data"].find({}).to_list(length=None)

    # client -> { payroll_month -> { employee_id -> {amount, travel_modes, entries} } }
    client_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {
        "amount": 0.0, "travel_modes": set(), "entries": []
    })))

    for odoc in all_ope_docs:
        emp_id = odoc.get("employeeId", "")
        emp_name = odoc.get("employeeName", "")
        for data_item in odoc.get("Data", []):
            for pm_key, elist in data_item.items():
                if payroll_month and pm_key != payroll_month:
                    continue
                for e in elist:
                    client = e.get("client", "Unknown") or "Unknown"
                    mode = (e.get("travel_mode") or "").lower()
                    amt = safe_float(e.get("amount", 0))

                    cd = client_data[client][pm_key][emp_id]
                    cd["amount"] += amt
                    cd["travel_modes"].add(mode)
                    cd["emp_name"] = emp_name
                    cd["entries"].append({
                        "date": e.get("date"),
                        "travel_mode": e.get("travel_mode"),
                        "amount": amt,
                        "project_id": e.get("project_id"),
                        "project_name": e.get("project_name"),
                    })

    # Build result
    pass_keywords = {"pass", "metro_pass", "bus_pass", "train_pass 1st", "train_pass 2nd"}
    ticket_keywords = {"ticket", "metro_tickets", "bus_ticket", "train_ticket"}

    result = []

    for client_name, months in client_data.items():
        client_summary = {
            "client": client_name,
            "payroll_months": [],
            "total_employees": 0,
            "total_amount": 0.0,
        }
        
        # Track unique employees across all months for this client
        unique_employees = set()

        for pm, emps in months.items():
            emp_list = []
            pm_total = 0.0
            has_pass_users = []
            has_ticket_users = []

            for eid, info in emps.items():
                unique_employees.add(eid)
                emp_amt = round(safe_float(info["amount"]), 2)
                pm_total += emp_amt
                modes = info["travel_modes"]

                uses_pass = bool(modes & pass_keywords)
                uses_ticket = bool(modes & ticket_keywords)

                emp_list.append({
                    "employee_id": eid,
                    "employee_name": info.get("emp_name", ""),
                    "amount": emp_amt,
                    "travel_modes": list(modes),
                    "uses_pass": uses_pass,
                    "uses_ticket": uses_ticket,
                    "entry_count": len(info["entries"]),
                })

                if uses_pass:
                    has_pass_users.append(eid)
                if uses_ticket:
                    has_ticket_users.append(eid)

            # Sort employees by amount descending
            emp_list.sort(key=lambda x: x["amount"], reverse=True)

            # Amount difference analysis
            if len(emp_list) >= 2:
                max_amt = emp_list[0]["amount"]
                min_amt = emp_list[-1]["amount"]
                amount_spread = round(safe_float(max_amt - min_amt), 2)

            else:
                amount_spread = 0.0

            # Pass vs ticket conflict
            pass_ticket_conflict = bool(has_pass_users and has_ticket_users)
            conflicting_pass_emps = [e for e in emp_list if e["uses_pass"]]
            conflicting_ticket_emps = [e for e in emp_list if e["uses_ticket"]]

            client_summary["payroll_months"].append({
                "payroll_month": pm,
                "employee_count": len(emp_list),
                "total_amount": round(pm_total, 2),
                "amount_spread": amount_spread,
                "employees": emp_list,
                "pass_ticket_conflict": pass_ticket_conflict,
                "pass_users": conflicting_pass_emps,
                "ticket_users": conflicting_ticket_emps,
            })

            client_summary["total_amount"] += pm_total

        # Set total employees to unique count
        client_summary["total_employees"] = len(unique_employees)

        client_summary["total_amount"] = round(client_summary["total_amount"], 2)
        result.append(client_summary)

    # Sort clients by total amount descending
    result.sort(key=lambda x: x["total_amount"], reverse=True)

    return {"clients": result, "total_clients": len(result)}


# ── 3. EXCEL EXPORT: CLIENT-WISE ─────────────────────────────
@app.get("/api/admin/export/client-wise")
async def export_client_excel(
    payroll_month: str = Query(None),
    current_user: dict = Depends(get_current_user)
):
    await verify_admin(current_user)

    all_ope_docs = await db["OPE_data"].find({}).to_list(length=None)
    all_status = await db["Status"].find({}).to_list(length=None)

    # Build employee → partner map
    emp_details_list = await db["Employee_details"].find({}).to_list(length=None)
    emp_partner_map = {e.get("EmpID", ""): e.get("Partner", "Unknown") for e in emp_details_list}

    rows = []
    for odoc in all_ope_docs:
        emp_id = odoc.get("employeeId", "")
        emp_name = odoc.get("employeeName", "")
        partner = emp_partner_map.get(emp_id, "Unknown")
        for data_item in odoc.get("Data", []):
            for pm_key, elist in data_item.items():
                if payroll_month and pm_key != payroll_month:
                    continue
                for e in elist:
                    rows.append({
                        "Employee ID": emp_id,
                        "Employee Name": emp_name,
                        "Partner": partner,
                        "Payroll Month": pm_key,
                        "Client": e.get("client", ""),
                        "Project ID": e.get("project_id", ""),
                        "Project Name": e.get("project_name", ""),
                        "Project Type": e.get("project_type", ""),
                        "Date": e.get("date", ""),
                        "Travel From": e.get("location_from", ""),
                        "Travel To": e.get("location_to", ""),
                        "Travel Mode": e.get("travel_mode", ""),
                        "Amount (₹)": safe_float(e.get("amount", 0)),
                        "Remarks": e.get("remarks", ""),
                        "Status": e.get("status", ""),
                    })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Client Wise Claims"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill = PatternFill("solid", fgColor="EBF3FF")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    if rows:
        headers = list(rows[0].keys())
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for ri, row in enumerate(rows, 2):
            fill = alt_fill if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for ci, key in enumerate(headers, 1):
                cell = ws.cell(row=ri, column=ci, value=row[key])
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        # Auto-width
        for ci, h in enumerate(headers, 1):
            col_letter = get_column_letter(ci)
            max_len = max(len(str(h)), max((len(str(r[h])) for r in rows), default=0))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
    else:
        ws["A1"] = "No data found"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"client_wise_claims{'_' + payroll_month if payroll_month else ''}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ── 4. EXCEL EXPORT: PARTNER-WISE ────────────────────────────
@app.get("/api/admin/export/partner-wise")
async def export_partner_excel(
    payroll_month: str = Query(None),
    current_user: dict = Depends(get_current_user)
):
    await verify_admin(current_user)

    all_status = await db["Status"].find({}).to_list(length=None)
    emp_details_list = await db["Employee_details"].find({}).to_list(length=None)
    emp_partner_map = {e.get("EmpID", ""): e.get("Partner", "Unknown") for e in emp_details_list}
    emp_name_map = {e.get("EmpID", ""): e.get("Emp Name", "") for e in emp_details_list}

    rows = []
    for sdoc in all_status:
        emp_id = sdoc.get("employeeId", "")
        emp_name = sdoc.get("employeeName", "") or emp_name_map.get(emp_id, "")
        partner = emp_partner_map.get(emp_id, "Unknown")

        for ps in sdoc.get("approval_status", []):
            pm = ps.get("payroll_month") or ps.get("month_range", "")
            if payroll_month and pm != payroll_month:
                continue
            rows.append({
                "Partner": partner,
                "Employee ID": emp_id,
                "Employee Name": emp_name,
                "Payroll Month": pm,
                "Total Amount (₹)": safe_float(ps.get("total_amount", 0)),
"OPE Limit (₹)": safe_float(ps.get("limit", 0)),
                "OPE Label": ps.get("ope_label", ""),
                "Overall Status": ps.get("overall_status", ""),
                "Current Level": ps.get("current_level", ""),
                "Total Levels": ps.get("total_levels", ""),
            })

    rows.sort(key=lambda x: (x["Partner"], x["Employee ID"]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Partner Wise Claims"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Group by partner for coloring
    partner_colors = {}
    color_list = ["EBF3FF", "FFF3E0", "E8F5E9", "FCE4EC", "F3E5F5", "E0F7FA"]
    for i, r in enumerate(rows):
        p = r["Partner"]
        if p not in partner_colors:
            partner_colors[p] = color_list[len(partner_colors) % len(color_list)]

    if rows:
        headers = list(rows[0].keys())
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for ri, row in enumerate(rows, 2):
            color = partner_colors.get(row["Partner"], "FFFFFF")
            fill = PatternFill("solid", fgColor=color)
            for ci, key in enumerate(headers, 1):
                cell = ws.cell(row=ri, column=ci, value=row[key])
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        for ci, h in enumerate(headers, 1):
            col_letter = get_column_letter(ci)
            max_len = max(len(str(h)), max((len(str(r[h])) for r in rows), default=0))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
    else:
        ws["A1"] = "No data found"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"partner_wise_claims{'_' + payroll_month if payroll_month else ''}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ── 5. DUPLICATE LOCATION TRAVEL CLAIMS DETECTION ───────────────
@app.get("/api/admin/analysis/duplicate-locations")
async def detect_duplicate_location_claims(
    payroll_month: str = Query(None),
    current_user: dict = Depends(get_current_user)
):
    """
    Detect duplicate location travel claims and short/excess amounts.
    Groups claims by location_from + location_to and identifies:
    - Duplicate claims (same route by same employee)
    - Amount variations (short/excess compared to average)
    """
    await verify_admin(current_user)
    
    all_ope_docs = await db["OPE_data"].find({}).to_list(length=None)
    
    # location_key -> list of claims
    location_claims = defaultdict(list)
    
    for odoc in all_ope_docs:
        emp_id = odoc.get("employeeId", "")
        emp_name = odoc.get("employeeName", "")
        for data_item in odoc.get("Data", []):
            for pm_key, elist in data_item.items():
                if payroll_month and pm_key != payroll_month:
                    continue
                for e in elist:
                    loc_from = (e.get("location_from") or "").strip().lower()
                    loc_to = (e.get("location_to") or "").strip().lower()
                    if loc_from and loc_to:
                        location_key = f"{loc_from}->{loc_to}"
                        location_claims[location_key].append({
                            "employee_id": emp_id,
                            "employee_name": emp_name,
                            "payroll_month": pm_key,
                            "date": e.get("date"),
                            "travel_mode": e.get("travel_mode"),
                            "amount": safe_float(e.get("amount", 0)),
                            "client": e.get("client"),
                            "project_name": e.get("project_name"),
                            "entry_id": str(e.get("_id", ""))
                        })
    
    # Analyze each location for duplicates and amount variations
    result = []
    for location_key, claims in location_claims.items():
        if len(claims) < 2:
            continue  # Skip if only one claim for this route
        
        # Group by employee to find duplicates
        emp_claims = defaultdict(list)
        for claim in claims:
            emp_claims[claim["employee_id"]].append(claim)
        
        # Find duplicates (same employee claiming same route multiple times)
        duplicates = []
        for emp_id, emp_claim_list in emp_claims.items():
            if len(emp_claim_list) > 1:
                duplicates.extend(emp_claim_list)
        
        # Calculate amount statistics
        amounts = [c["amount"] for c in claims]
        avg_amount = sum(amounts) / len(amounts) if amounts else 0
        min_amount = min(amounts) if amounts else 0
        max_amount = max(amounts) if amounts else 0
        
        # Identify short/excess claims
        short_claims = [c for c in claims if c["amount"] < avg_amount * 0.8]
        excess_claims = [c for c in claims if c["amount"] > avg_amount * 1.2]
        
        result.append({
            "location_route": location_key,
            "total_claims": len(claims),
            "unique_employees": len(emp_claims),
            "duplicate_claims": len(duplicates),
            "duplicate_details": duplicates,
            "amount_stats": {
                "average": round(avg_amount, 2),
                "minimum": round(min_amount, 2),
                "maximum": round(max_amount, 2),
                "spread": round(max_amount - min_amount, 2)
            },
            "short_claims": short_claims,
            "excess_claims": excess_claims,
            "all_claims": claims
        })
    
    # Sort by duplicate count descending
    result.sort(key=lambda x: x["duplicate_claims"], reverse=True)
    
    return {
        "location_analysis": result,
        "total_locations_analyzed": len(result),
        "total_duplicate_claims": sum(r["duplicate_claims"] for r in result)
    }


# ── 6. BACKDATED OPE CLAIMS ANALYSIS ────────────────────────────
@app.get("/api/admin/analysis/backdated-claims")
async def analyze_backdated_claims(
    days_threshold: int = Query(30, description="Days after payroll month to consider as backdated"),
    current_user: dict = Depends(get_current_user)
):
    """
    Analyze all backdated OPE claims.
    Identifies claims submitted after the payroll month end date + threshold days.
    """
    await verify_admin(current_user)
    
    all_ope_docs = await db["OPE_data"].find({}).to_list(length=None)
    backdated_claims = []
    
    for odoc in all_ope_docs:
        emp_id = odoc.get("employeeId", "")
        emp_name = odoc.get("employeeName", "")
        for data_item in odoc.get("Data", []):
            for pm_key, elist in data_item.items():
                # Parse payroll month (format: "Dec 2025 - Jan 2026" or "Jan 2026")
                try:
                    # Extract the end month from payroll month range
                    if "-" in pm_key:
                        parts = pm_key.split("-")
                        end_month_str = parts[-1].strip()
                    else:
                        end_month_str = pm_key.strip()
                    
                    # Parse the month year (e.g., "Jan 2026")
                    month_year = end_month_str.split()
                    if len(month_year) >= 2:
                        month_name = month_year[0]
                        year = int(month_year[1])
                        
                        # Map month names to numbers
                        month_map = {
                            "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
                            "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12
                        }
                        month_num = month_map.get(month_name.lower(), 1)
                        
                        # Calculate end of payroll month (last day of the month)
                        last_day = calendar.monthrange(year, month_num)[1]
                        payroll_end_date = datetime(year, month_num, last_day)
                        threshold_date = payroll_end_date + timedelta(days=days_threshold)
                    else:
                        continue
                except Exception as e:
                    print(f"Error parsing payroll month {pm_key}: {e}")
                    continue
                
                for e in elist:
                    submission_date = e.get("submission_date") or e.get("created_time")
                    if submission_date:
                        if isinstance(submission_date, str):
                            try:
                                submission_date = datetime.fromisoformat(submission_date.replace('Z', '+00:00'))
                            except:
                                continue
                        elif isinstance(submission_date, datetime):
                            pass
                        else:
                            continue
                        
                        # Check if submission is after threshold
                        if submission_date > threshold_date:
                            backdated_claims.append({
                                "employee_id": emp_id,
                                "employee_name": emp_name,
                                "payroll_month": pm_key,
                                "payroll_end_date": payroll_end_date.isoformat(),
                                "threshold_date": threshold_date.isoformat(),
                                "submission_date": submission_date.isoformat(),
                                "days_late": (submission_date - threshold_date).days,
                                "date": e.get("date"),
                                "client": e.get("client"),
                                "project_name": e.get("project_name"),
                                "amount": safe_float(e.get("amount", 0)),
                                "travel_mode": e.get("travel_mode"),
                                "location_from": e.get("location_from"),
                                "location_to": e.get("location_to")
                            })
    
    # Sort by days late descending
    backdated_claims.sort(key=lambda x: x["days_late"], reverse=True)
    
    return {
        "backdated_claims": backdated_claims,
        "total_backdated_claims": len(backdated_claims),
        "threshold_days": days_threshold
    }


# ── 7. PROJECT-WISE CONSOLIDATED OPE REPORTS ─────────────────────
@app.get("/api/admin/analysis/project-wise")
async def project_wise_consolidated_reports(
    payroll_month: str = Query(None),
    current_user: dict = Depends(get_current_user)
):
    """
    Generate project-wise and monthly consolidated OPE reports.
    Aggregates data by project and shows monthly breakdowns.
    """
    await verify_admin(current_user)
    
    all_ope_docs = await db["OPE_data"].find({}).to_list(length=None)
    
    # project -> { payroll_month -> { employee_id -> amount } }
    project_data = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    project_details = {}
    
    for odoc in all_ope_docs:
        emp_id = odoc.get("employeeId", "")
        emp_name = odoc.get("employeeName", "")
        for data_item in odoc.get("Data", []):
            for pm_key, elist in data_item.items():
                if payroll_month and pm_key != payroll_month:
                    continue
                for e in elist:
                    project_id = e.get("project_id", "Unknown")
                    project_name = e.get("project_name", "Unknown")
                    project_type = e.get("project_type", "Unknown")
                    client = e.get("client", "Unknown")
                    amt = safe_float(e.get("amount", 0))
                    
                    project_data[project_id][pm_key][emp_id] += amt
                    
                    # Store project details
                    if project_id not in project_details:
                        project_details[project_id] = {
                            "project_id": project_id,
                            "project_name": project_name,
                            "project_type": project_type,
                            "client": client
                        }
    
    # Build consolidated report
    result = []
    for project_id, months in project_data.items():
        project_summary = {
            **project_details[project_id],
            "payroll_months": [],
            "total_employees": 0,
            "total_amount": 0.0
        }
        
        unique_employees = set()
        
        for pm, emps in months.items():
            pm_total = sum(emps.values())
            unique_employees.update(emps.keys())
            
            project_summary["payroll_months"].append({
                "payroll_month": pm,
                "employee_count": len(emps),
                "total_amount": round(pm_total, 2),
                "employees": [{"employee_id": eid, "amount": round(amt, 2)} for eid, amt in emps.items()]
            })
            
            project_summary["total_amount"] += pm_total
        
        project_summary["total_employees"] = len(unique_employees)
        project_summary["total_amount"] = round(project_summary["total_amount"], 2)
        result.append(project_summary)
    
    # Sort by total amount descending
    result.sort(key=lambda x: x["total_amount"], reverse=True)
    
    return {
        "projects": result,
        "total_projects": len(result)
    }


# ── 8. PROJECT-WISE ANALYSIS FOR EACH CLIENT ─────────────────────
@app.get("/api/admin/analysis/client-project-wise")
async def client_project_wise_analysis(
    payroll_month: str = Query(None),
    current_user: dict = Depends(get_current_user)
):
    """
    Project-wise analysis for each client.
    Shows which projects are more/less profitable within each client,
    and relates to employee count.
    """
    await verify_admin(current_user)
    
    all_ope_docs = await db["OPE_data"].find({}).to_list(length=None)
    
    # client -> project -> { payroll_month -> { employee_id -> amount } }
    client_project_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(float))))
    project_details = {}
    
    for odoc in all_ope_docs:
        emp_id = odoc.get("employeeId", "")
        emp_name = odoc.get("employeeName", "")
        for data_item in odoc.get("Data", []):
            for pm_key, elist in data_item.items():
                if payroll_month and pm_key != payroll_month:
                    continue
                for e in elist:
                    client = e.get("client", "Unknown")
                    project_id = e.get("project_id", "Unknown")
                    project_name = e.get("project_name", "Unknown")
                    project_type = e.get("project_type", "Unknown")
                    amt = safe_float(e.get("amount", 0))
                    
                    client_project_data[client][project_id][pm_key][emp_id] += amt
                    
                    # Store project details
                    key = f"{client}_{project_id}"
                    if key not in project_details:
                        project_details[key] = {
                            "project_id": project_id,
                            "project_name": project_name,
                            "project_type": project_type,
                            "client": client
                        }
    
    # Build analysis result
    result = []
    
    # REPLACE this block (around line with "for client, projects in client_project_data.items():")
    for client, projects in client_project_data.items():
        client_summary = {
            "client": client,
            "projects": [],
            "total_employees": 0,
            "total_amount": 0.0,
            "most_profitable_project": None,
            "least_profitable_project": None
        }
        
        client_unique_employees = set()
        project_summaries = []
        
        for project_id, months in projects.items():
            key = f"{client}_{project_id}"
            project_summary = {
                **project_details[key],
                "payroll_months": [],
                "total_employees": 0,
                "total_amount": 0.0,
                "avg_amount_per_employee": 0.0
            }
            
            project_unique_employees = set()
            
            for pm, emps in months.items():
                pm_total = sum(emps.values())
                project_unique_employees.update(emps.keys())
                client_unique_employees.update(emps.keys())
                
                project_summary["payroll_months"].append({
                    "payroll_month": pm,
                    "employee_count": len(emps),
                    "total_amount": round(pm_total, 2)
                })
                
                project_summary["total_amount"] += pm_total
            
            project_summary["total_employees"] = len(project_unique_employees)
            project_summary["total_amount"] = round(project_summary["total_amount"], 2)
            
            if project_summary["total_employees"] > 0:
                project_summary["avg_amount_per_employee"] = round(
                    project_summary["total_amount"] / project_summary["total_employees"], 2
                )
            
            client_summary["total_amount"] += project_summary["total_amount"]
            project_summaries.append(project_summary)
        
        client_summary["total_employees"] = len(client_unique_employees)
        client_summary["total_amount"] = round(client_summary["total_amount"], 2)
        
        # Sort projects by amount descending
        project_summaries.sort(key=lambda x: x["total_amount"], reverse=True)
        client_summary["projects"] = project_summaries
        
        # Identify most and least profitable projects
        if project_summaries:
            client_summary["most_profitable_project"] = project_summaries[0]
            client_summary["least_profitable_project"] = project_summaries[-1]
        
        result.append(client_summary)
    
    # Sort clients by total amount descending
    result.sort(key=lambda x: x["total_amount"], reverse=True)
    
    # At the end of client_project_wise_analysis, before "return":
    # Compute globally unique project IDs and employee IDs
    all_unique_project_ids = set()
    all_unique_emp_ids = set()
    for client, projects in client_project_data.items():
        for project_id, months in projects.items():
            all_unique_project_ids.add(project_id)
            for pm, emps in months.items():
                all_unique_emp_ids.update(emps.keys())

        return {
            "clients": result,
            "total_clients": len(result),
            "global_unique_projects": len(all_unique_project_ids),   # ADD THIS
            "global_unique_employees": len(all_unique_emp_ids),       # ADD THIS
        }
    # return {
    #     "clients": result,
    #     "total_clients": len(result)
    # }
    
@app.get("/api/projects/{employee_code}")
async def get_employee_projects(employee_code: str, current_user=Depends(get_current_user)):
    """
    Fetch projects from Timesheets.Projects filtered by the
    employee's partner_emp_code.
    Returns a flat project list + a deduplicated client list.
    """
    try:
        emp_code = employee_code.strip().upper()
 
        if current_user["employee_code"].upper() != emp_code:
            raise HTTPException(status_code=403, detail="Access denied")
 
        # Get employee's partner code from Employee_details
        emp = await db["Employee_details"].find_one({"EmpID": emp_code})
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
 
        partner_emp_code = emp.get("PartnerEmpCode", "").strip().upper()
        print(f"📂 Projects request — emp: {emp_code}, partner: {partner_emp_code}")
 
        if not partner_emp_code:
            return {"projects": [], "clients": []}
 
        # Query the Timesheets database (different DB on same cluster)
        timesheets_db = client["Timesheets"]
        cursor = timesheets_db["Projects"].find(
            {"partner_emp_code": partner_emp_code},
            {"_id": 0}      # exclude ObjectId to keep JSON simple
        )
        raw = await cursor.to_list(length=None)
        print(f"✅ Found {len(raw)} projects for partner {partner_emp_code}")
 
        # Deduplicate by project_code
        seen = set()
        projects = []
        for p in raw:
            code = (p.get("project_code") or "").strip()
            if code in seen:
                continue
            seen.add(code)
            projects.append({
                "project_code": code,
                "project_name": (p.get("project_name") or "").strip(),
                "client_code":  (p.get("client_code")  or "").strip(),
                "client_name":  (p.get("client_name")  or "").strip(),
            })
 
        # Build sorted unique client list
        client_map = {}
        for p in projects:
            if p["client_code"] and p["client_name"]:
                client_map[p["client_code"]] = p["client_name"]
 
        clients = [
            {"client_code": k, "client_name": v}
            for k, v in sorted(client_map.items(), key=lambda x: x[1].lower())
        ]
 
        return {
            "projects": projects,
            "clients":  clients,
        }
 
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ get_employee_projects error: {e}")
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# ---------- Serve static HTML ----------
app.mount("/", StaticFiles(directory="static", html=True), name="static")

